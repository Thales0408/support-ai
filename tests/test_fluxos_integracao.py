import io
import unittest
from contextlib import contextmanager
from unittest.mock import patch

from openpyxl import load_workbook

import app
from services import ai


class FakeCursor:

    def __init__(self, state):

        self.state = state
        self.result = None

    def __enter__(self):

        return self

    def __exit__(self, exc_type, exc, tb):

        return False

    def execute(self, sql, params=None):

        params = params or ()
        sql_lower = " ".join(str(sql).lower().split())
        self.state["queries"].append((sql_lower, params))

        if "from usuarios" in sql_lower and "where usuario = %s" in sql_lower:

            usuario = params[0]
            user = self.state["users_by_name"].get(usuario)
            self.result = [
                (
                    user["id"],
                    user["senha"],
                    user["is_admin"],
                    user["ativo"],
                    user["perfil"]
                )
            ] if user else []
            return

        if "insert into atendimentos" in sql_lower and "returning id" in sql_lower:

            atendimento_id = self.state["next_atendimento_id"]
            self.state["next_atendimento_id"] += 1
            self.state["atendimentos"][atendimento_id] = {
                "id": atendimento_id,
                "usuario_id": params[0],
                "arquivo": params[1],
                "conteudo": params[2],
                "data": params[3],
                "status": params[4],
                "ticket_zendesk": params[5],
                "transcricao_completa": "",
                "duracao_segundos": 0,
                "chunks_total": 0,
                "chunks_falhos": 0,
                "chunks_ignorados": 0,
                "segundos_transcritos": 0,
                "custo_estimado_usd": 0,
                "resumo_editado": False,
                "sentimento_cliente": "neutro",
                "urgencia": "media",
                "categoria": "outro",
                "problema_principal": "",
                "tags": ""
            }
            self.result = [(atendimento_id,)]
            return

        if "select 1 from atendimentos" in sql_lower:

            atendimento_id = int(params[0])
            usuario_id = int(params[1])
            atendimento = self.state["atendimentos"].get(atendimento_id)
            self.result = [(1,)] if atendimento and atendimento["usuario_id"] == usuario_id else []
            return

        if "select status, provider_usado, fallback_usado from transcricoes_chunks" in sql_lower:

            key = (int(params[0]), int(params[2]))
            chunk = self.state["chunks"].get(key)
            self.result = [
                (
                    chunk["status"],
                    chunk.get("provider_usado"),
                    chunk.get("fallback_usado")
                )
            ] if chunk else []
            return

        if "insert into transcricoes_chunks" in sql_lower:

            key = (int(params[0]), int(params[2]))
            self.state["chunks"][key] = {
                "atendimento_id": int(params[0]),
                "usuario_id": int(params[1]),
                "ordem": int(params[2]),
                "texto": params[3],
                "status": "transcrito",
                "provider_tentado": params[4],
                "provider_usado": params[5],
                "fallback_usado": bool(params[6]),
                "duracao_segundos": int(params[7])
            }
            self.result = []
            return

        if "update atendimentos set status = 'transcrevendo'" in sql_lower:

            atendimento = self.state["atendimentos"].get(int(params[0]))
            if atendimento:
                atendimento["status"] = "transcrevendo"
            self.result = []
            return

        if "update atendimentos set status = 'finalizando'" in sql_lower:

            atendimento = self.state["atendimentos"].get(int(params[0]))
            if atendimento:
                atendimento["status"] = "finalizando"
            self.result = []
            return

        if "select status, conteudo, chunks_total" in sql_lower and "for update" in sql_lower:

            atendimento = self.state["atendimentos"].get(int(params[0]))
            usuario_id = int(params[1])
            if not atendimento or atendimento["usuario_id"] != usuario_id:
                self.result = []
                return
            self.result = [(
                atendimento["status"],
                atendimento["conteudo"],
                atendimento["chunks_total"],
                atendimento["chunks_falhos"],
                atendimento["chunks_ignorados"],
                atendimento["segundos_transcritos"],
                atendimento["custo_estimado_usd"]
            )]
            return

        if "select texto, status from transcricoes_chunks" in sql_lower:

            atendimento_id = int(params[0])
            usuario_id = int(params[1])
            chunks = [
                chunk
                for chunk in self.state["chunks"].values()
                if chunk["atendimento_id"] == atendimento_id
                and chunk["usuario_id"] == usuario_id
            ]
            self.result = [
                (
                    chunk["texto"],
                    chunk["status"]
                )
                for chunk in sorted(chunks, key=lambda item: item["ordem"])
            ]
            return

        if "select count(*), sum(" in sql_lower and "from transcricoes_chunks" in sql_lower:

            atendimento_id = int(params[0])
            usuario_id = int(params[1])
            chunks = [
                chunk
                for chunk in self.state["chunks"].values()
                if chunk["atendimento_id"] == atendimento_id
                and chunk["usuario_id"] == usuario_id
            ]
            self.result = [(
                len(chunks),
                len([chunk for chunk in chunks if chunk["status"] == "erro"])
            )]
            return

        if "select coalesce(provider_usado" in sql_lower:

            atendimento_id = int(params[2])
            usuario_id = int(params[3])
            chunks = [
                chunk
                for chunk in self.state["chunks"].values()
                if chunk["atendimento_id"] == atendimento_id
                and chunk["usuario_id"] == usuario_id
                and chunk["status"] == "transcrito"
            ]
            self.result = [
                (
                    chunk["provider_usado"],
                    chunk["duracao_segundos"]
                )
                for chunk in sorted(chunks, key=lambda item: item["ordem"])
            ]
            return

        if "update atendimentos set conteudo = %s" in sql_lower and "status = 'finalizado'" in sql_lower:

            atendimento = self.state["atendimentos"].get(int(params[13]))
            if atendimento:
                atendimento.update({
                    "conteudo": params[0],
                    "transcricao_completa": params[1],
                    "status": "finalizado",
                    "duracao_segundos": params[2],
                    "chunks_total": params[3],
                    "chunks_falhos": params[4],
                    "chunks_ignorados": params[5],
                    "segundos_transcritos": params[6],
                    "custo_estimado_usd": params[7],
                    "sentimento_cliente": params[8],
                    "urgencia": params[9],
                    "categoria": params[10],
                    "problema_principal": params[11],
                    "tags": params[12]
                })
            self.result = []
            return

        if "insert into uso_eventos" in sql_lower:

            self.state["uso_eventos"].append(params)
            self.result = []
            return

        if "from atendimentos a left join usuarios u" in sql_lower and "limit 300" in sql_lower:

            rows = []
            for atendimento in sorted(
                self.state["atendimentos"].values(),
                key=lambda item: item["id"],
                reverse=True
            ):
                rows.append(self._linha_resultado(atendimento))
            self.result = rows
            return

        if "select id, usuario from usuarios" in sql_lower:

            self.result = [
                (user["id"], user["usuario"])
                for user in self.state["users_by_name"].values()
                if user["ativo"]
            ]
            return

        if "where a.id = %s and a.usuario_id = %s" in sql_lower:

            atendimento = self.state["atendimentos"].get(int(params[0]))
            if not atendimento or atendimento["usuario_id"] != int(params[1]):
                self.result = []
                return
            self.result = [self._linha_detalhe(atendimento)]
            return

        if "where a.id = %s" in sql_lower and "from atendimentos a" in sql_lower:

            atendimento = self.state["atendimentos"].get(int(params[0]))
            self.result = [self._linha_detalhe(atendimento)] if atendimento else []
            return

        if "from atendimentos a" in sql_lower and "order by a.id desc" in sql_lower:

            self.result = [
                self._linha_exportacao(atendimento)
                for atendimento in sorted(
                    self.state["atendimentos"].values(),
                    key=lambda item: item["id"],
                    reverse=True
                )
            ]
            return

        self.result = []

    def _usuario_nome(self, usuario_id):

        for user in self.state["users_by_name"].values():
            if user["id"] == usuario_id:
                return user["usuario"]
        return ""

    def _linha_resultado(self, atendimento):

        return (
            atendimento["id"],
            atendimento["arquivo"],
            atendimento["conteudo"],
            atendimento["data"],
            atendimento["status"],
            atendimento["chunks_total"],
            atendimento["chunks_falhos"],
            atendimento["duracao_segundos"],
            atendimento["transcricao_completa"],
            atendimento["ticket_zendesk"],
            atendimento["chunks_ignorados"],
            atendimento["segundos_transcritos"],
            atendimento["custo_estimado_usd"],
            atendimento["resumo_editado"],
            atendimento["sentimento_cliente"],
            atendimento["urgencia"],
            atendimento["categoria"],
            atendimento["problema_principal"],
            atendimento["tags"],
            self._usuario_nome(atendimento["usuario_id"]),
            atendimento["usuario_id"]
        )

    def _linha_detalhe(self, atendimento):

        return (
            atendimento["id"],
            atendimento["conteudo"],
            atendimento["transcricao_completa"],
            atendimento["data"],
            atendimento["status"],
            atendimento["duracao_segundos"],
            atendimento["chunks_total"],
            atendimento["chunks_falhos"],
            atendimento["ticket_zendesk"],
            atendimento["chunks_ignorados"],
            atendimento["segundos_transcritos"],
            atendimento["custo_estimado_usd"],
            atendimento["resumo_editado"],
            atendimento["sentimento_cliente"],
            atendimento["urgencia"],
            atendimento["categoria"],
            atendimento["problema_principal"],
            atendimento["tags"],
            self._usuario_nome(atendimento["usuario_id"])
        )

    def _linha_exportacao(self, atendimento):

        return (
            atendimento["data"],
            self._usuario_nome(atendimento["usuario_id"]),
            atendimento["ticket_zendesk"],
            atendimento["conteudo"],
            atendimento["transcricao_completa"],
            atendimento["chunks_total"],
            atendimento["chunks_falhos"],
            atendimento["chunks_ignorados"],
            atendimento["segundos_transcritos"],
            atendimento["custo_estimado_usd"],
            atendimento["sentimento_cliente"],
            atendimento["urgencia"],
            atendimento["categoria"],
            atendimento["problema_principal"],
            atendimento["tags"]
        )

    def fetchone(self):

        if not self.result:
            return None
        return self.result[0]

    def fetchall(self):

        return list(self.result or [])


class FakeConn:

    def __init__(self, state):

        self.state = state

    def __enter__(self):

        return self

    def __exit__(self, exc_type, exc, tb):

        return False

    def cursor(self):

        return FakeCursor(self.state)


class FluxosIntegracaoTest(unittest.TestCase):

    def setUp(self):

        app.app.config.update(TESTING=True)
        self.state = {
            "next_atendimento_id": 10,
            "users_by_name": {
                "analista": {
                    "id": 1,
                    "usuario": "analista",
                    "senha": "senha",
                    "is_admin": False,
                    "ativo": True,
                    "perfil": "analista"
                },
                "supervisor": {
                    "id": 2,
                    "usuario": "supervisor",
                    "senha": "senha",
                    "is_admin": False,
                    "ativo": True,
                    "perfil": "supervisor"
                },
                "admin": {
                    "id": 3,
                    "usuario": "admin",
                    "senha": "senha",
                    "is_admin": True,
                    "ativo": True,
                    "perfil": "admin_tecnico"
                }
            },
            "atendimentos": {
                99: {
                    "id": 99,
                    "usuario_id": 2,
                    "arquivo": "streaming",
                    "conteudo": app.resumo_zendesk_exato(
                        nome_empresa="Empresa X",
                        analista="supervisor",
                        descritivo="Atendimento finalizado."
                    ),
                    "data": "03/06/2026 09:00",
                    "status": "finalizado",
                    "ticket_zendesk": "ZD-99",
                    "transcricao_completa": "Cliente solicitou suporte.",
                    "duracao_segundos": 60,
                    "chunks_total": 1,
                    "chunks_falhos": 0,
                    "chunks_ignorados": 0,
                    "segundos_transcritos": 60,
                    "custo_estimado_usd": 0.05,
                    "resumo_editado": False,
                    "sentimento_cliente": "neutro",
                    "urgencia": "media",
                    "categoria": "fiscal",
                    "problema_principal": "NFS-e",
                    "tags": "fiscal,nfse"
                }
            },
            "chunks": {},
            "uso_eventos": [],
            "queries": []
        }

        self.patchers = [
            patch("app.conectar_banco", self.fake_connect),
            patch("app.login_bloqueado", return_value=False),
            patch("app.registrar_login_falho", return_value=None),
            patch("app.limpar_tentativas_login", return_value=None),
            patch("app.log_evento", return_value=None),
            patch("app.uso_diario_usuario", return_value={
                "chamadas": 0,
                "segundos": 0,
                "chunks": 0,
                "custo": 0
            }),
            patch("app.validar_limites_custo_resumo", return_value=None),
            patch("app.validar_limite_custo_fallback_transcricao", return_value=None)
        ]
        for patcher in self.patchers:
            patcher.start()

        self.client = app.app.test_client()

    def tearDown(self):

        for patcher in reversed(self.patchers):
            patcher.stop()

    @contextmanager
    def fake_connect(self):

        yield FakeConn(self.state)

    def set_session(self, usuario_id=1, perfil="analista", nome="analista"):

        with self.client.session_transaction() as sess:
            sess["usuario_id"] = usuario_id
            sess["perfil"] = perfil
            sess["is_admin"] = perfil == "admin_tecnico"
            sess["usuario_nome"] = nome
            sess["csrf_token"] = "csrf-teste"

    def post_json(self, url, payload=None, token_header="X-CSRFToken"):

        return self.client.post(
            url,
            json=payload or {},
            headers={token_header: "csrf-teste"}
        )

    def test_login_logout_dashboard(self):

        with self.client.session_transaction() as sess:
            sess["csrf_token"] = "csrf-teste"

        response = self.client.post(
            "/login",
            data={
                "usuario": "analista",
                "senha": "senha",
                "csrf_token": "csrf-teste"
            },
            follow_redirects=False
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers["Location"], "/")

        dashboard = self.client.get("/")
        self.assertEqual(dashboard.status_code, 200)
        self.assertIn(b"55PBX AI", dashboard.data)

        logout = self.client.get("/logout", follow_redirects=False)
        self.assertEqual(logout.status_code, 302)
        self.assertIn("/login", logout.headers["Location"])

    def test_csrf_exigido_e_aceita_header_sem_hifen(self):

        self.set_session()

        sem_token = self.client.post("/atendimentos/iniciar", json={})
        self.assertEqual(sem_token.status_code, 403)

        com_token = self.post_json("/atendimentos/iniciar")
        self.assertEqual(com_token.status_code, 200)
        self.assertEqual(com_token.get_json()["status"], "gravando")

    def test_fluxo_gravacao_chunk_finalizacao_zendesk(self):

        self.set_session(nome="analista")

        inicio = self.post_json(
            "/atendimentos/iniciar",
            {"ticket_zendesk": "ZD-10"}
        )
        self.assertEqual(inicio.status_code, 200)
        atendimento_id = inicio.get_json()["atendimento_id"]

        with patch("app.transcrever_chunk", return_value={
            "texto": (
                "Cliente informou CNPJ 08 633 889 mil contra 56 "
                "e suporte equipamentos arroba gmail ponto com"
            ),
            "provider_tentado": "groq",
            "provider_usado": "groq",
            "fallback_usado": False
        }):
            chunk = self.client.post(
                "/atendimentos/chunk",
                data={
                    "atendimento_id": str(atendimento_id),
                    "ordem": "0",
                    "duracao_segundos": "30",
                    "audio": (
                        io.BytesIO(b"a" * 2048),
                        "chunk.webm"
                    )
                },
                headers={"X-CSRFToken": "csrf-teste"},
                content_type="multipart/form-data"
            )

        self.assertEqual(chunk.status_code, 200)
        self.assertEqual(chunk.get_json()["status"], "chunk_transcrito")
        self.assertIn("E-mail identificado", chunk.get_json()["texto"])

        with patch("app.analisar_com_ia", return_value={
            "resumo_zendesk": app.resumo_zendesk_exato(
                nome_empresa="Empresa Teste",
                cnpj="Possivel CNPJ informado: 08.633.889/0001-56 - confirmar com cliente",
                email="suporteequipamentos@gmail.com",
                analista="analista",
                descritivo="Cliente solicitou analise de cadastro fiscal."
            ),
            "sentimento_cliente": "neutro",
            "urgencia": "media",
            "categoria": "fiscal",
            "problema_principal": "Cadastro fiscal",
            "tags": "fiscal,cadastro"
        }) as mock_ia:
            finalizar = self.post_json(
                "/atendimentos/finalizar",
                {
                    "atendimento_id": atendimento_id,
                    "duracao_segundos": 30,
                    "chunks_total": 1,
                    "chunks_falhos": 0,
                    "chunks_ignorados": 0,
                    "segundos_transcritos": 30
                }
            )

        self.assertEqual(finalizar.status_code, 200)
        resposta = finalizar.get_json()
        self.assertEqual(resposta["status"], "finalizado")
        self.assertIn("Nome da empresa:", resposta["resultado"])
        self.assertIn("E-mail Solicitante: suporteequipamentos@gmail.com", resposta["resultado"])
        self.assertEqual(mock_ia.call_count, 1)

        with patch("app.analisar_com_ia") as mock_ia_repetida:
            repetida = self.post_json(
                "/atendimentos/finalizar",
                {"atendimento_id": atendimento_id}
            )

        self.assertEqual(repetida.status_code, 200)
        self.assertTrue(repetida.get_json()["reutilizado"])
        self.assertEqual(mock_ia_repetida.call_count, 0)

    def test_permissoes_custos_dashboard_detalhe_exportacao(self):

        self.set_session(usuario_id=1, perfil="analista", nome="analista")
        detalhe_analista = self.client.get("/atendimentos/99")
        self.assertEqual(detalhe_analista.status_code, 404)

        self.set_session(usuario_id=2, perfil="supervisor", nome="supervisor")
        resultados_supervisor = self.client.get("/resultados")
        self.assertEqual(resultados_supervisor.status_code, 200)
        dados_supervisor = resultados_supervisor.get_json()
        self.assertTrue(dados_supervisor["is_supervisor"])
        self.assertFalse(dados_supervisor["mostrar_custo"])
        self.assertNotIn("usd_brl_rate", dados_supervisor)
        self.assertNotIn("custo_estimado_usd", dados_supervisor["resultados"][0])

        detalhe_supervisor = self.client.get("/atendimentos/99")
        self.assertEqual(detalhe_supervisor.status_code, 200)
        self.assertNotIn("custo_estimado_usd", detalhe_supervisor.get_json())

        export_supervisor = self.client.get("/exportar")
        self.assertEqual(export_supervisor.status_code, 200)
        workbook_supervisor = load_workbook(
            io.BytesIO(export_supervisor.data)
        )
        cabecalho_supervisor = [
            cell.value
            for cell in workbook_supervisor.active[1]
        ]
        self.assertNotIn("Custo estimado USD", cabecalho_supervisor)
        self.assertNotIn("Custo estimado BRL", cabecalho_supervisor)
        workbook_supervisor.close()

        self.set_session(usuario_id=3, perfil="admin_tecnico", nome="admin")
        resultados_admin = self.client.get("/resultados")
        self.assertEqual(resultados_admin.status_code, 200)
        dados_admin = resultados_admin.get_json()
        self.assertTrue(dados_admin["mostrar_custo"])
        self.assertIn("usd_brl_rate", dados_admin)
        self.assertIn("custo_estimado_usd", dados_admin["resultados"][0])

        detalhe_admin = self.client.get("/atendimentos/99")
        self.assertEqual(detalhe_admin.status_code, 200)
        self.assertIn("custo_estimado_usd", detalhe_admin.get_json())

        export_admin = self.client.get("/exportar")
        self.assertEqual(export_admin.status_code, 200)
        workbook_admin = load_workbook(
            io.BytesIO(export_admin.data)
        )
        cabecalho_admin = [
            cell.value
            for cell in workbook_admin.active[1]
        ]
        self.assertIn("Custo estimado USD", cabecalho_admin)
        self.assertIn("Custo estimado BRL", cabecalho_admin)
        workbook_admin.close()

    def test_finalizacao_libera_status_em_limite_de_duracao(self):

        self.set_session()
        self.state["atendimentos"][10] = {
            "id": 10,
            "usuario_id": 1,
            "arquivo": "streaming",
            "conteudo": "Transcricao em andamento...",
            "data": "03/06/2026 09:00",
            "status": "transcrevendo",
            "ticket_zendesk": "",
            "transcricao_completa": "",
            "duracao_segundos": 0,
            "chunks_total": 0,
            "chunks_falhos": 0,
            "chunks_ignorados": 0,
            "segundos_transcritos": 0,
            "custo_estimado_usd": 0,
            "resumo_editado": False,
            "sentimento_cliente": "neutro",
            "urgencia": "media",
            "categoria": "outro",
            "problema_principal": "",
            "tags": ""
        }

        response = self.post_json(
            "/atendimentos/finalizar",
            {
                "atendimento_id": 10,
                "duracao_segundos": app.MAX_CALL_DURATION_MINUTES * 60 + 1
            }
        )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(self.state["atendimentos"][10]["status"], "transcrevendo")


class FallbackTranscricaoTest(unittest.TestCase):

    def test_fallback_groq_para_openai(self):

        chamadas = []

        def fake_transcrever_bytes(provider, audio_bytes, nome, mime):

            chamadas.append(provider)
            if provider == "groq":
                raise RuntimeError("rate limit reached")
            return "texto via openai"

        arquivo = io.BytesIO(b"audio")
        arquivo.filename = "chunk.webm"
        arquivo.mimetype = "audio/webm"

        with patch("services.ai.TRANSCRIBE_PROVIDER", "groq"), \
                patch("services.ai.TRANSCRIBE_FALLBACK_PROVIDER", "openai"), \
                patch("services.ai.transcrever_bytes", fake_transcrever_bytes):
            resultado = ai.transcrever_chunk(arquivo)

        self.assertEqual(chamadas, ["groq", "openai"])
        self.assertEqual(resultado["provider_tentado"], "groq")
        self.assertEqual(resultado["provider_usado"], "openai")
        self.assertTrue(resultado["fallback_usado"])
        self.assertEqual(resultado["texto"], "texto via openai")


if __name__ == "__main__":

    unittest.main()
