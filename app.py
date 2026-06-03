from flask import (
    Flask,
    request,
    jsonify,
    render_template,
    redirect,
    session,
    send_file,
    abort
)

from flask_cors import CORS
from openai import RateLimitError
from waitress import serve
from datetime import datetime
from openpyxl import Workbook
from werkzeug.security import generate_password_hash

from io import BytesIO
import os
import psycopg2
import re
import uuid
import traceback
import json
import logging
import unicodedata
import secrets
import hmac

from auth import (
    perfil_usuario,
    perfil_valido,
    senha_esta_em_hash,
    senha_valida,
    usuario_admin_tecnico,
    usuario_logado,
    usuario_supervisor
)
from config import (
    CHUNK_SECONDS,
    CORS_ORIGINS,
    LOGIN_BLOCK_MINUTES,
    LOGIN_MAX_ATTEMPTS,
    MAX_AUDIO_MINUTES_PER_DAY,
    MAX_CALL_DURATION_MINUTES,
    MAX_CALLS_PER_DAY,
    MAX_CHUNKS_PER_CALL,
    SECRET_KEY,
    SUMMARY_MODEL,
    TRANSCRIBE_FALLBACK_PROVIDER,
    TRANSCRIBE_PROVIDER,
    USD_BRL_RATE,
)
from services.ai import (
    LimiteCustoFallbackTranscricao,
    cliente_resumo,
    estimar_custo_atendimento,
    estimar_custo_transcricao,
    estimar_custo_transcricao_por_provedor,
    transcrever_chunk
)
from services.database import (
    conectar_banco,
    inicializar_banco
)
from services.usage import (
    avaliar_limites_custo_transcricao,
    avaliar_limites_custo_resumo,
    custo_brl,
    registrar_uso_evento,
    uso_diario_usuario
)


# =========================================
# FLASK
# =========================================

logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO"),
    format="%(asctime)s %(levelname)s %(message)s"
)

logger = logging.getLogger("support_ai")

app = Flask(__name__)

app.secret_key = SECRET_KEY

if CORS_ORIGINS:

    CORS(
        app,
        origins=CORS_ORIGINS,
        supports_credentials=True
    )


def obter_csrf_token():

    token = session.get("csrf_token")

    if not token:

        token = secrets.token_urlsafe(32)
        session["csrf_token"] = token

    return token


@app.context_processor
def contexto_csrf():

    return {
        "csrf_token": obter_csrf_token
    }


@app.before_request
def proteger_csrf():

    if request.method != "POST":

        return None

    esperado = session.get("csrf_token")
    recebido = (
        request.headers.get("X-CSRF-Token")
        or request.headers.get("X-CSRFToken")
        or request.form.get("csrf_token")
    )

    if (
        not esperado
        or not recebido
        or not hmac.compare_digest(str(esperado), str(recebido))
    ):

        if request.path.startswith((
            "/atendimentos",
            "/conta"
        )):

            return jsonify({
                "erro": "Token CSRF invalido"
            }), 403

        abort(403)

    return None


# =========================================
# STARTUP
# =========================================

try:

    inicializar_banco()

except Exception as e:

    print("ERRO AO INICIALIZAR BANCO:", e)


# =========================================
# HELPERS
# =========================================

def log_evento(evento, **dados):

    logger.info(
        json.dumps(
            {
                "evento": evento,
                **dados
            },
            ensure_ascii=False,
            default=str
        )
    )


def limpar_texto(texto):

    texto = re.sub(
        r"\s+",
        " ",
        texto or ""
    )

    return texto.strip()


def limpar_vazamento_prompt_transcricao(texto):

    texto_limpo = str(texto or "")
    frases_prompt = [
        "Transcreva em português do Brasil",
        "Transcreva em portugues do Brasil",
        "Texto e atendimento de suporte",
        "Atendimento de suporte ERP",
        "Contexto: atendimento de suporte tecnico",
        "Contexto: atendimento de suporte técnico",
        "Não invente palavras quando houver silêncio",
        "Nao invente palavras quando houver silencio",
        "Não invente palavras quando houver silencio",
        "Nao invente palavras quando houver silêncio"
    ]

    for frase in frases_prompt:

        texto_limpo = re.sub(
            re.escape(frase),
            "",
            texto_limpo,
            flags=re.IGNORECASE
        )

    texto_limpo = re.sub(
        r"\s+",
        " ",
        texto_limpo
    )

    return texto_limpo.strip(" .:-")


def tamanho_arquivo_upload(arquivo):

    posicao_atual = (
        arquivo.stream.tell()
    )

    arquivo.stream.seek(
        0,
        os.SEEK_END
    )

    tamanho = (
        arquivo.stream.tell()
    )

    arquivo.stream.seek(
        posicao_atual
    )

    return tamanho


def ip_requisicao():

    encaminhado = request.headers.get("X-Forwarded-For", "")

    if encaminhado:

        return encaminhado.split(",")[0].strip()

    return request.remote_addr or "desconhecido"


def login_bloqueado(usuario, ip):

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                """
                SELECT tentativas, bloqueado_ate
                FROM login_tentativas
                WHERE usuario = %s
                AND ip = %s
                """,
                (
                    usuario,
                    ip
                )
            )

            row = cursor.fetchone()

            if not row:

                return False

            tentativas, bloqueado_ate = row

            if bloqueado_ate:

                cursor.execute(
                    "SELECT NOW() < %s",
                    (
                        bloqueado_ate,
                    )
                )

                if cursor.fetchone()[0]:

                    log_evento(
                        "login_bloqueado",
                        usuario=usuario,
                        ip=ip,
                        tentativas=tentativas,
                        bloqueado_ate=bloqueado_ate
                    )

                    return True

    return False


def registrar_login_falho(usuario, ip):

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                """
                INSERT INTO login_tentativas (
                    usuario,
                    ip,
                    tentativas,
                    bloqueado_ate,
                    atualizado_em
                )
                VALUES (%s, %s, 1, NULL, NOW())
                ON CONFLICT (usuario, ip)
                DO UPDATE SET
                    tentativas = login_tentativas.tentativas + 1,
                    bloqueado_ate = CASE
                        WHEN login_tentativas.tentativas + 1 >= %s
                        THEN NOW() + (%s || ' minutes')::interval
                        ELSE login_tentativas.bloqueado_ate
                    END,
                    atualizado_em = NOW()
                RETURNING tentativas, bloqueado_ate
                """,
                (
                    usuario,
                    ip,
                    LOGIN_MAX_ATTEMPTS,
                    LOGIN_BLOCK_MINUTES
                )
            )

            tentativas, bloqueado_ate = cursor.fetchone()

    log_evento(
        "login_falhou",
        usuario=usuario,
        ip=ip,
        tentativas=tentativas,
        bloqueado_ate=bloqueado_ate
    )


def limpar_tentativas_login(usuario, ip):

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                """
                DELETE FROM login_tentativas
                WHERE usuario = %s
                AND ip = %s
                """,
                (
                    usuario,
                    ip
                )
            )


def validar_limites_custo_resumo(
    cursor,
    usuario_id,
    atendimento_id,
    custo_estimado_usd
):

    limite = avaliar_limites_custo_resumo(
        cursor,
        usuario_id,
        atendimento_id,
        custo_estimado_usd
    )

    if not limite:

        return None

    log_evento(
        limite["evento"],
        **limite["log"]
    )

    return erro_limite(
        limite["mensagem"],
        tipo=limite["evento"],
        deve_parar_gravacao=True,
        **limite["resposta"]
    )


def validar_limite_custo_fallback_transcricao(
    cursor,
    usuario_id,
    atendimento_id,
    custo_estimado_usd
):

    limite = avaliar_limites_custo_transcricao(
        cursor,
        usuario_id,
        atendimento_id,
        custo_estimado_usd
    )

    if not limite:

        return None

    log_evento(
        limite["evento"],
        **limite["log"]
    )

    return erro_limite(
        limite["mensagem"],
        tipo=limite["evento"],
        deve_parar_gravacao=True,
        **limite["resposta"]
    )


def usuario_filtro_atendimentos():

    usuario_id = usuario_logado()
    analista = request.args.get("analista_id") or request.args.get("escopo")

    if not usuario_supervisor():

        return {
            "where": "a.usuario_id = %s",
            "params": [usuario_id],
            "escopo": "meus",
            "analista_id": usuario_id
        }

    if analista in [
        "todos",
        "all"
    ]:

        return {
            "where": "1 = 1",
            "params": [],
            "escopo": "todos",
            "analista_id": None
        }

    if (
        not analista
        or analista == "meus"
    ):

        return {
            "where": "a.usuario_id = %s",
            "params": [usuario_id],
            "escopo": "meus",
            "analista_id": usuario_id
        }

    try:

        analista_id = int(analista)

    except (TypeError, ValueError):

        analista_id = usuario_id

    return {
        "where": "a.usuario_id = %s",
        "params": [analista_id],
        "escopo": "analista",
        "analista_id": analista_id
    }


def erro_limite(
    mensagem,
    tipo="limite_uso",
    deve_parar_gravacao=True,
    **dados
):

    return jsonify({
        "erro": "limite_atingido",
        "limite": True,
        "tipo": tipo,
        "mensagem": mensagem,
        "deve_parar_gravacao": deve_parar_gravacao,
        **dados
    }), 403


def registrar_erro_chunk(
    usuario_id,
    atendimento_id,
    ordem_int,
    erro,
    status_atendimento="erro_chunk"
):

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                """
                INSERT INTO transcricoes_chunks (
                    atendimento_id,
                    usuario_id,
                    ordem,
                    texto,
                    status,
                    erro,
                    provider_tentado
                )
                VALUES (%s, %s, %s, %s, 'erro', %s, %s)
                ON CONFLICT (atendimento_id, ordem)
                DO UPDATE SET
                    texto = EXCLUDED.texto,
                    status = 'erro',
                    erro = EXCLUDED.erro,
                    provider_tentado = EXCLUDED.provider_tentado
                """,
                (
                    atendimento_id,
                    usuario_id,
                    ordem_int,
                    "",
                    str(erro)[:500],
                    TRANSCRIBE_PROVIDER
                )
            )

            cursor.execute(
                """
                UPDATE atendimentos
                SET status = %s
                WHERE id = %s
                AND usuario_id = %s
                """,
                (
                    status_atendimento,
                    atendimento_id,
                    usuario_id
                )
            )


def erro_rate_limit_transcricao(e):

    status_code = getattr(e, "status_code", None)
    texto = str(e).lower()

    return (
        status_code == 429
        or "rate limit" in texto
        or "too many requests" in texto
        or "requests per day" in texto
    )


def calcular_segundos_por_provider(chunks_transcritos, segundos_transcritos):

    segundos_restantes = max(0, int(segundos_transcritos or 0))
    segundos_por_provider = {}

    for provider, duracao_chunk in chunks_transcritos:

        if segundos_restantes <= 0:

            break

        segundos_chunk = min(
            max(1, int(duracao_chunk or CHUNK_SECONDS)),
            segundos_restantes
        )
        provider_normalizado = provider or TRANSCRIBE_PROVIDER

        segundos_por_provider[provider_normalizado] = (
            segundos_por_provider.get(provider_normalizado, 0)
            + segundos_chunk
        )
        segundos_restantes -= segundos_chunk

    return segundos_por_provider


def extrair_json_objeto(texto):

    bruto = (
        str(texto or "").strip()
    )

    if bruto.startswith("```"):

        bruto = (
            re.sub(
                r"^```(?:json)?",
                "",
                bruto,
                flags=re.IGNORECASE
            ).strip()
        )

        bruto = (
            re.sub(
                r"```$",
                "",
                bruto
            ).strip()
        )

    inicio = (
        bruto.find("{")
    )

    fim = (
        bruto.rfind("}")
    )

    if inicio >= 0 and fim > inicio:

        bruto = (
            bruto[inicio:fim + 1]
        )

    return json.loads(bruto)


def normalizar_campo(valor, padrao="", limite=300):

    texto = (
        limpar_texto(valor or "")
    )

    if not texto:

        texto = padrao

    return texto[:limite]


def normalizar_campo_zendesk(valor, padrao="", limite=300):

    texto = normalizar_campo(
        valor,
        padrao,
        limite
    )
    texto = remover_frases_proibidas_zendesk(texto)

    if not texto:

        return padrao

    return texto


def remover_acentos(texto):

    return "".join(
        caractere
        for caractere in unicodedata.normalize(
            "NFD",
            str(texto or "")
        )
        if unicodedata.category(caractere) != "Mn"
    )


FRASES_PROIBIDAS_ZENDESK = [
    "nao " + "informado",
    "nao informada",
    "nao identificado",
    "nao identificada",
    "resumo nao disponivel pela transcricao",
    "transcricao confusa",
    "nao ha informacoes suficientes",
    "nao foi possivel identificar",
    "nao apresentou conclusao clara",
    "a conclusao nao ficou clara"
]


ROTULOS_ZENDESK = [
    "Nome da empresa",
    "Empresa/Loja",
    "CNPJ",
    "Nome do Cliente",
    "Telefone de contato",
    "E-mail Solicitante",
    "Email Solicitante",
    "Analista responsavel",
    "Analista responsável",
    "Descritivo da ocorrencia do atendimento",
    "Descritivo da ocorrência do atendimento"
]


def normalizar_para_comparacao(texto):

    return remover_acentos(
        str(texto or "")
    ).lower().strip()


def remover_frases_proibidas_zendesk(texto):

    resultado = str(texto or "")
    padroes = [
        r"N[aã]o informado\.?",
        r"N[aã]o informada\.?",
        r"N[aã]o identificado(?: na liga[cç][aã]o)?\.?",
        r"N[aã]o identificada(?: na liga[cç][aã]o)?\.?",
        r"Resumo n[aã]o dispon[ií]vel pela transcri[cç][aã]o\.?",
        r"A transcri[cç][aã]o est[aá] confusa\.?",
        r"Transcri[cç][aã]o confusa\.?",
        r"N[aã]o h[aá] informa[cç][oõ]es suficientes\.?",
        r"N[aã]o foi poss[ií]vel identificar\.?",
        r"O atendimento n[aã]o apresentou conclus[aã]o clara\.?",
        r"A conclus[aã]o n[aã]o ficou clara\.?"
    ]

    for padrao in padroes:

        resultado = re.sub(
            padrao,
            "",
            resultado,
            flags=re.IGNORECASE
        )

    comparacao = normalizar_para_comparacao(resultado)

    if any(frase in comparacao for frase in FRASES_PROIBIDAS_ZENDESK):

        return ""

    return resultado.strip()


def validar_cnpj_digitos(digitos):

    if (
        len(digitos) != 14
        or len(set(digitos)) == 1
    ):

        return False

    pesos_primeiro = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    pesos_segundo = [6] + pesos_primeiro

    def calcular(posicoes, pesos):

        soma = sum(
            int(digito) * peso
            for digito, peso in zip(posicoes, pesos)
        )
        resto = soma % 11

        return "0" if resto < 2 else str(11 - resto)

    primeiro = calcular(digitos[:12], pesos_primeiro)
    segundo = calcular(digitos[:12] + primeiro, pesos_segundo)

    return digitos[-2:] == primeiro + segundo


def formatar_cnpj(digitos):

    return (
        f"{digitos[0:2]}.{digitos[2:5]}.{digitos[5:8]}/"
        f"{digitos[8:12]}-{digitos[12:14]}"
    )


def normalizar_cnpj(valor, permitir_possivel=False):

    texto = str(valor or "")

    if texto.lower().startswith("possível cnpj informado:"):

        return texto[:120]

    digitos = re.sub(
        r"\D",
        "",
        texto
    )

    if len(digitos) == 14 and validar_cnpj_digitos(digitos):

        return formatar_cnpj(digitos)

    if permitir_possivel and len(digitos) == 14:

        return (
            "Possível CNPJ informado: "
            + formatar_cnpj(digitos)
            + " — confirmar com cliente"
        )

    return ""


def possivel_cnpj_formatado(digitos):

    return (
        "Possível CNPJ informado: "
        + formatar_cnpj(digitos)
        + " — confirmar com cliente"
    )


PALAVRAS_NUMERO_CNPJ = {
    "zero": "0",
    "um": "1",
    "uma": "1",
    "hum": "1",
    "dois": "2",
    "duas": "2",
    "tres": "3",
    "três": "3",
    "quatro": "4",
    "cinco": "5",
    "seis": "6",
    "meia": "6",
    "sete": "7",
    "oito": "8",
    "nove": "9"
}


PALAVRAS_DEZENA_CNPJ = {
    "dez": 10,
    "onze": 11,
    "doze": 12,
    "treze": 13,
    "quatorze": 14,
    "catorze": 14,
    "quinze": 15,
    "dezesseis": 16,
    "dezessete": 17,
    "dezoito": 18,
    "dezenove": 19,
    "vinte": 20,
    "trinta": 30,
    "quarenta": 40,
    "cinquenta": 50,
    "sessenta": 60,
    "setenta": 70,
    "oitenta": 80,
    "noventa": 90
}


def valor_numero_cnpj(token):

    token_limpo = remover_acentos(token)

    if token_limpo.isdigit():

        valor = int(token_limpo)

        if 1 <= valor <= 9:

            return valor

    if token_limpo in PALAVRAS_NUMERO_CNPJ:

        return int(PALAVRAS_NUMERO_CNPJ[token_limpo])

    return None


def contar_digitos_cnpj_simples(texto):

    total = 0
    tokens = re.findall(
        r"\d+|[^\W\d_]+",
        str(texto or "").lower()
    )

    for token in tokens:

        token_limpo = remover_acentos(token)

        if token.isdigit():

            total += len(token)

        elif token_limpo in PALAVRAS_NUMERO_CNPJ:

            total += 1

        elif token_limpo in PALAVRAS_DEZENA_CNPJ:

            total += 2

    return total


def normalizar_blocos_cnpj_falados(texto):

    texto_normalizado = str(texto or "")

    def bloco(prefixo):

        numero = valor_numero_cnpj(prefixo or "um") or 1

        return f"000{numero}"

    def substituir_posfixo(match):

        prefixo = match.group(1)

        if prefixo:

            digitos_antes = contar_digitos_cnpj_simples(
                texto_normalizado[:match.start(1)]
            )

            if digitos_antes in [6, 7]:

                return prefixo + " 0001"

        return bloco(prefixo)

    padrao_posfixo = re.compile(
        (
            r"\b(?:(um|uma|hum|dois|duas|tres|tr[eê]s|quatro|cinco|"
            r"seis|sete|oito|nove|[1-9])\s+)?mil\s+"
            r"(?:contra|contr[aá]rio|de\s+r[eé]|dere|dre|barra)\b"
        ),
        flags=re.IGNORECASE
    )

    texto_normalizado = padrao_posfixo.sub(
        substituir_posfixo,
        texto_normalizado
    )

    padrao_barra = re.compile(
        (
            r"\bbarra\s+(?:(um|uma|hum|dois|duas|tres|tr[eê]s|quatro|"
            r"cinco|seis|sete|oito|nove|[1-9])\s+)?mil\b"
        ),
        flags=re.IGNORECASE
    )

    return padrao_barra.sub(
        lambda match: bloco(match.group(1)),
        texto_normalizado
    )


def tem_bloco_cnpj_falado(texto):

    texto_normalizado = normalizar_para_comparacao(texto)

    return bool(
        re.search(
            (
                r"(?:^|\s)(?:(?:um|uma|hum|dois|duas|tres|quatro|cinco|"
                r"seis|sete|oito|nove|[1-9])\s+)?mil\s+"
                r"(?:contra|contrario|de re|dere|dre|barra)(?:\s|$)"
                r"|(?:^|\s)barra\s+(?:(?:um|uma|hum|dois|duas|tres|quatro|"
                r"cinco|seis|sete|oito|nove|[1-9])\s+)?mil(?:\s|$)"
            ),
            texto_normalizado
        )
    )


def texto_para_digitos_cnpj(texto):

    texto = normalizar_blocos_cnpj_falados(texto)

    tokens = re.findall(
        r"\d+|[^\W\d_]+",
        str(texto or "").lower()
    )
    partes = []
    indice = 0

    while indice < len(tokens):

        token = tokens[indice]
        token_limpo = remover_acentos(token)

        if token.isdigit():

            partes.append(token)
            indice += 1
            continue

        if token_limpo in PALAVRAS_NUMERO_CNPJ:

            partes.append(PALAVRAS_NUMERO_CNPJ[token_limpo])
            indice += 1
            continue

        if token_limpo in PALAVRAS_DEZENA_CNPJ:

            valor = PALAVRAS_DEZENA_CNPJ[token_limpo]

            if (
                indice + 2 < len(tokens)
                and remover_acentos(tokens[indice + 1]) == "e"
                and remover_acentos(tokens[indice + 2])
                in PALAVRAS_NUMERO_CNPJ
            ):

                valor += int(
                    PALAVRAS_NUMERO_CNPJ[
                        remover_acentos(tokens[indice + 2])
                    ]
                )
                indice += 3

            else:

                indice += 1

            partes.append(str(valor).zfill(2))
            continue

        indice += 1

    return "".join(partes)


def grupos_numericos_cnpj(texto):

    return re.findall(
        r"\d+",
        str(texto or "")
    )


def candidatos_cnpj_por_grupos(fragmento):

    inferido_contexto = tem_bloco_cnpj_falado(fragmento)
    fragmento = normalizar_blocos_cnpj_falados(fragmento)
    grupos = grupos_numericos_cnpj(fragmento)
    candidatos = []

    if len(grupos) >= 5:

        base = "".join(grupos[:3])
        bloco = grupos[3]
        final = grupos[4]

        if (
            len(base) == 8
            and len(bloco) == 4
            and len(final) >= 2
        ):

            if re.fullmatch(r"[1-9]000", bloco):

                candidatos.append(
                    (
                        base + "000" + bloco[0] + final[-2:],
                        True
                    )
                )

            if bloco == "1000" and len(final) >= 3:

                candidatos.append(
                    (
                        base + "0001" + final[-2:],
                        True
                    )
                )
                candidatos.append(
                    (
                        base + bloco + final[:2],
                        True
                    )
                )

            candidatos.append(
                (
                    base + bloco + final[-2:],
                    inferido_contexto
                )
            )

    return [
        candidato
        for candidato in candidatos
        if len(candidato[0]) == 14
    ]


def fragmentos_com_evidencia_cnpj(texto):

    texto = str(texto or "")
    evidencias = [
        "cnpj",
        "c n p j",
        "cadastro nacional",
        "barra",
        "traco",
        "traço",
        "contrario",
        "contrário",
        "contra",
        "de re",
        "de rÃ©",
        "dere",
        "dre",
        "mil"
    ]

    for match in re.finditer(
        "|".join(re.escape(evidencia) for evidencia in evidencias),
        texto,
        flags=re.IGNORECASE
    ):

        inicio = max(0, match.start() - 80)
        fim = min(len(texto), match.end() + 140)

        yield texto[inicio:fim]

    for match in re.finditer(
        r"\d[\d\s.,/\-]{12,}\d",
        texto
    ):

        inicio = max(0, match.start() - 30)
        fim = min(len(texto), match.end() + 30)

        yield texto[inicio:fim]


def extrair_possivel_cnpj(texto):

    melhor_possivel = ""
    melhor_bruto = ""

    for fragmento in fragmentos_com_evidencia_cnpj(texto):

        digitos = texto_para_digitos_cnpj(fragmento)
        candidatos = candidatos_cnpj_por_grupos(fragmento)
        inferido_por_fala = tem_bloco_cnpj_falado(fragmento)

        for inicio in range(0, max(1, len(digitos) - 13)):

            candidato = digitos[inicio:inicio + 14]

            if len(candidato) == 14:

                candidatos.append((candidato, inferido_por_fala))

        if (
            len(digitos) == 13
            and digitos[8:11] == "000"
        ):

            candidatos.append(
                (
                    digitos[:11] + "1" + digitos[11:],
                    True
                )
            )

        if (
            len(digitos) == 13
            and inferido_por_fala
        ):

            candidatos.append(
                (
                    "0" + digitos,
                    True
                )
            )

        for candidato, inferido in dict.fromkeys(candidatos):

            if validar_cnpj_digitos(candidato):

                if inferido:

                    return possivel_cnpj_formatado(candidato)

                return formatar_cnpj(candidato)

            if not melhor_possivel:

                melhor_possivel = candidato

        if not melhor_bruto:

            bruto = re.search(
                r"\d[\d\s.,/\-]{7,}\d",
                fragmento
            )

            if bruto:

                valor_bruto = re.sub(
                    r"\s+",
                    " ",
                    bruto.group(0)
                ).strip(" ,.;:")
                digitos_bruto = re.sub(
                    r"\D",
                    "",
                    valor_bruto
                )

                if 8 <= len(digitos_bruto) < 14:

                    melhor_bruto = valor_bruto

    if melhor_possivel:

        return possivel_cnpj_formatado(melhor_possivel)

    if melhor_bruto:

        return (
            "Possível CNPJ informado: "
            + melhor_bruto
            + " — confirmar com cliente"
        )

    return ""


def normalizar_descritivo_zendesk(valor):

    texto = str(valor or "").strip()
    texto = remover_frases_proibidas_zendesk(texto)

    texto = re.sub(
        r"\*\*|__|`",
        "",
        texto
    )

    texto = re.sub(
        r"(?m)^\s*[-*]\s+",
        "",
        texto
    )

    texto = re.sub(
        r"[ \t]+",
        " ",
        texto
    )

    texto = re.sub(
        r"\n{3,}",
        "\n\n",
        texto
    ).strip()

    texto = re.sub(
        r"\s+([.!?])",
        r"\1",
        texto
    ).strip()

    if not texto:

        return ""

    if not campo_zendesk_informado(texto):

        return ""

    return texto[:1200]


def campo_zendesk_informado(valor):

    texto = limpar_texto(
        remover_frases_proibidas_zendesk(valor)
    )

    if not texto:

        return False

    comparacao = normalizar_para_comparacao(texto)

    return comparacao not in [
        "nao " + "informado",
        "nao informada",
        "nao identificado",
        "nao identificada",
        "null",
        "none",
        "n/a",
        "...",
        "-"
    ]


def valor_obrigatorio_zendesk(valor):

    texto = limpar_valor_estruturado(valor, limite=160)

    if campo_zendesk_informado(texto):

        return texto

    return ""


def limpar_valor_estruturado(valor, limite=160):

    texto = normalizar_campo_zendesk(
        valor,
        padrao="",
        limite=limite
    )
    texto = re.sub(
        r"\s+",
        " ",
        texto
    ).strip()
    texto = remover_frases_proibidas_zendesk(texto)

    comparacao = normalizar_para_comparacao(texto)

    if "possivel cnpj informado" in comparacao:

        return ""

    if re.search(
        r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}",
        texto
    ):

        return ""

    for rotulo in ROTULOS_ZENDESK:

        if re.fullmatch(
            re.escape(rotulo) + r"\s*:?",
            texto,
            flags=re.IGNORECASE
        ):

            return ""

        if re.match(
            r"^" + re.escape(rotulo) + r"\s*:",
            texto,
            flags=re.IGNORECASE
        ):

            return ""

        match_rotulo = re.search(
            r"\s+" + re.escape(rotulo) + r"\s*:",
            texto,
            flags=re.IGNORECASE
        )

        if match_rotulo:

            texto = texto[:match_rotulo.start()].strip()

    if re.fullmatch(r"[A-Za-zÀ-ÿ /-]{2,45}\s*:", texto):

        return ""

    return texto[:limite]


def normalizar_email_zendesk(valor):

    texto = limpar_valor_estruturado(
        valor,
        limite=120
    )

    if not campo_zendesk_informado(texto):

        return ""

    match = re.search(
        r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}",
        texto,
        flags=re.IGNORECASE
    )

    if not match:

        return normalizar_email_falado(texto)

    email = match.group(0)
    dominio = email.split("@", 1)[1]

    if (
        not dominio
        or "." not in dominio
        or dominio.startswith(".")
        or dominio.endswith(".")
    ):

        return ""

    return email


def montar_parte_email_falado(tokens):

    partes = []

    for token in tokens:

        token = remover_acentos(token).lower()

        if token == "ponto":

            partes.append(".")

        elif token in [
            "underline",
            "sublinhado"
        ]:

            partes.append("_")

        elif token in [
            "traco",
            "hifen"
        ]:

            partes.append("-")

        elif token in [
            "e",
            "email",
            "e-mail"
        ]:

            continue

        else:

            partes.append(
                re.sub(
                    r"[^a-z0-9]",
                    "",
                    token
                )
            )

    return "".join(partes).strip(".-_")


def normalizar_email_falado(texto):

    texto_original = str(texto or "")
    match_formatado = re.search(
        r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}",
        texto_original,
        flags=re.IGNORECASE
    )

    if match_formatado:

        return match_formatado.group(0).lower()

    texto_normalizado = normalizar_para_comparacao(texto_original)
    tokens = re.findall(
        r"[a-z0-9]+",
        texto_normalizado
    )

    if "arroba" not in tokens:

        return ""

    indice = tokens.index("arroba")
    palavras_parada = {
        "meu",
        "minha",
        "email",
        "e",
        "mail",
        "solicitante",
        "cliente",
        "favor",
        "por",
        "para"
    }

    antes = []

    for token in reversed(tokens[:indice]):

        if token in palavras_parada and antes:

            break

        if token in palavras_parada:

            continue

        antes.append(token)

        if len(antes) >= 6:

            break

    local = montar_parte_email_falado(
        list(reversed(antes))
    )

    depois = []

    for token in tokens[indice + 1:]:

        if token in palavras_parada and depois:

            break

        if token in palavras_parada:

            continue

        depois.append(token)

        if len(depois) >= 8:

            break

    dominio = montar_parte_email_falado(depois)
    candidato = f"{local}@{dominio}".lower()

    if re.fullmatch(
        r"[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}",
        candidato
    ):

        return candidato

    return ""


def normalizar_entidades_faladas(texto):

    texto_base = str(texto or "").strip()

    if not texto_base:

        return ""

    return normalizar_blocos_cnpj_falados(texto_base)


def limpar_nome_participante(valor):

    texto = limpar_texto(valor or "")
    texto = re.split(
        (
            r"\b(?:tudo bem|bom dia|boa tarde|boa noite|como vai|"
            r"suporte|atendimento|gestao|gestão|click)\b"
        ),
        texto,
        maxsplit=1,
        flags=re.IGNORECASE
    )[0]
    texto = re.sub(
        r"^(?:o|a|eu sou|me chamo)\s+",
        "",
        texto,
        flags=re.IGNORECASE
    )
    texto = re.sub(
        r"[^A-Za-zÀ-ÿ\s'-]",
        "",
        texto
    )
    texto = re.sub(
        r"\s+",
        " ",
        texto
    ).strip()

    partes = texto.split()

    if not partes or len(partes) > 4:

        return ""

    if any(len(parte) <= 1 for parte in partes):

        return ""

    return " ".join(
        parte[:1].upper() + parte[1:]
        for parte in partes
    )


def nomes_iguais(nome_a, nome_b):

    return (
        bool(nome_a)
        and bool(nome_b)
        and normalizar_para_comparacao(nome_a)
        == normalizar_para_comparacao(nome_b)
    )


def extrair_analista_nome(texto):

    padroes = [
        r"\bmeu nome (?:é|e)\s+([^,.;\n]{2,60})",
        r"\bsou (?:o|a)?\s*([^,.;\n]{2,60})",
        r"\bfala com\s+([^,.;\n]{2,60})",
        r"\baqui (?:é|e)\s+([^,.;\n]{2,60})"
    ]

    for padrao in padroes:

        match = re.search(
            padrao,
            str(texto or ""),
            flags=re.IGNORECASE
        )

        if match:

            nome = limpar_nome_participante(
                match.group(1)
            )

            if nome:

                return nome

    return ""


def extrair_cliente_nome(texto, analista_nome=""):

    padroes = [
        r"\bnome do cliente (?:é|e)\s+([^,.;\n]{2,60})",
        r"\bcliente se chama\s+([^,.;\n]{2,60})",
        r"\bcliente[:\s]+([^,.;\n]{2,60})"
    ]

    for padrao in padroes:

        match = re.search(
            padrao,
            str(texto or ""),
            flags=re.IGNORECASE
        )

        if match:

            nome = limpar_nome_participante(
                match.group(1)
            )

            if nome and not nomes_iguais(nome, analista_nome):

                return nome

    return ""


def extrair_empresa_transcricao(texto):

    padroes = [
        r"\braz[aã]o social (?:é|e|da|do)?\s+([^,.;\n]{3,100})",
        r"\bempresa (?:é|e|da|do)?\s+([^,.;\n]{3,100})",
        r"\bloja (?:é|e|da|do)?\s+([^,.;\n]{3,100})"
    ]

    for padrao in padroes:

        match = re.search(
            padrao,
            str(texto or ""),
            flags=re.IGNORECASE
        )

        if match:

            empresa = limpar_valor_estruturado(
                match.group(1),
                limite=120
            )

            if campo_zendesk_informado(empresa):

                return empresa

    return ""


def extrair_telefone_transcricao(texto):

    match = re.search(
        r"(?:telefone|celular|whats(?:app)?)\D{0,20}(\+?\d[\d\s().-]{7,}\d)",
        str(texto or ""),
        flags=re.IGNORECASE
    )

    if not match:

        return ""

    return limpar_valor_estruturado(
        match.group(1),
        limite=40
    )


def limpar_transcricao_para_resumo(texto):

    texto = limpar_vazamento_prompt_transcricao(
        str(texto or "")
    )
    texto = re.sub(
        r"\s+",
        " ",
        texto
    ).strip()

    if not texto:

        return ""

    ruidos = {
        "nis",
        "hum",
        "uh",
        "ahn",
        "hã",
        "ha",
        "e aí",
        "e ai"
    }

    frases = re.split(
        r"(?<=[.!?])\s+|(?:\s+-\s+)",
        texto
    )
    limpas = []

    for frase in frases:

        frase = frase.strip(" ,;:")

        if not frase:

            continue

        comparacao = normalizar_para_comparacao(frase)

        if comparacao in ruidos:

            continue

        if (
            len(comparacao) <= 3
            and not re.search(r"\d|@", frase)
        ):

            continue

        limpas.append(frase)

    texto = " ".join(limpas)

    texto = re.sub(
        r"\b(al[oô]\s*){3,}",
        "alô ",
        texto,
        flags=re.IGNORECASE
    )
    texto = re.sub(
        r"\b(nis[, ]+){2,}nis\b",
        "",
        texto,
        flags=re.IGNORECASE
    )
    texto = re.sub(
        r"\b(\w{2,})(?:\s+\1){2,}\b",
        r"\1",
        texto,
        flags=re.IGNORECASE
    )

    return re.sub(
        r"\s+",
        " ",
        texto
    ).strip()


def extrair_entidades_transcricao(texto):

    texto_base = str(texto or "").strip()
    analista_nome = extrair_analista_nome(texto_base)

    return {
        "analista_nome": analista_nome,
        "cliente_nome": extrair_cliente_nome(
            texto_base,
            analista_nome
        ),
        "empresa": extrair_empresa_transcricao(texto_base),
        "cnpj": extrair_possivel_cnpj(texto_base),
        "email": normalizar_email_falado(texto_base),
        "telefone": extrair_telefone_transcricao(texto_base)
    }


def texto_entidades_extraidas(entidades):

    entidades = entidades or {}
    linhas = []

    for chave in [
        "analista_nome",
        "cliente_nome",
        "empresa",
        "telefone"
    ]:

        if entidades.get(chave):

            linhas.append(
                chave + ": " + entidades[chave]
            )

    if entidades.get("cnpj"):

        linhas.append(
            "cnpj: " + entidades["cnpj"]
        )

    if entidades.get("email"):

        linhas.append(
            "email: " + entidades["email"]
        )

    return "\n".join(linhas)


def remover_rotulos_do_descritivo(valor):

    texto = str(valor or "")
    rotulos = [
        "Nome da empresa",
        "Empresa/Loja",
        "CNPJ",
        "Nome do Cliente",
        "Telefone de contato",
        "E-mail Solicitante",
        "Email Solicitante",
        "Analista responsável",
        "Analista responsavel",
        "Descritivo da ocorrência do atendimento",
        "Descritivo da ocorrencia do atendimento"
    ]

    for rotulo in rotulos:

        texto = re.sub(
            r"(?im)^\s*" + re.escape(rotulo) + r"\s*:.*$",
            "",
            texto
        )

    return texto.strip()


def resumo_zendesk_exato(
    nome_empresa=None,
    empresa_loja=None,
    cnpj=None,
    cnpj_contexto=None,
    cliente=None,
    telefone=None,
    email=None,
    analista=None,
    descritivo=None
):

    cnpj_final = normalizar_cnpj(cnpj, permitir_possivel=True)

    if not cnpj_final:

        cnpj_final = extrair_possivel_cnpj(
            " ".join([
                str(cnpj or ""),
                str(cnpj_contexto or "")
            ])
        )

    if not campo_zendesk_informado(cnpj_final):

        cnpj_final = ""

    campos = [
        "Nome da empresa: " + valor_obrigatorio_zendesk(nome_empresa),
        "Empresa/Loja: " + valor_obrigatorio_zendesk(empresa_loja),
        "CNPJ: " + cnpj_final,
        "Nome do Cliente: " + valor_obrigatorio_zendesk(cliente),
        "Telefone de contato: " + valor_obrigatorio_zendesk(telefone),
        "E-mail Solicitante: " + normalizar_email_zendesk(email)
    ]

    campos.extend([
        "Analista responsável: " + normalizar_campo_zendesk(analista, limite=120),
        (
            "Descritivo da ocorrência do atendimento:\n"
            + normalizar_descritivo_zendesk(
                remover_rotulos_do_descritivo(descritivo)
            )
        )
    ])

    return "\n\n".join(campos)


def extrair_secao_texto(texto, rotulos):

    for rotulo in rotulos:

        padrao = (
            r"(?:^|\n)\s*"
            + re.escape(rotulo)
            + r"\s*:\s*(.*?)(?=\n\s*[A-Za-zÀ-ÿ][A-Za-zÀ-ÿ /-]{1,45}\s*:|\Z)"
        )

        match = re.search(
            padrao,
            texto,
            flags=re.IGNORECASE | re.DOTALL
        )

        if match:

            return limpar_texto(
                match.group(1)
            )

    return ""


def texto_zendesk_formatado(conteudo):

    texto = (
        str(conteudo or "").strip()
    )

    if not texto:

        return ""

    texto_extracao = re.sub(
        (
            r"\s+("
            r"Empresa/Loja:|"
            r"CNPJ:|"
            r"Nome do Cliente:|"
            r"Telefone de contato:|"
            r"E-mail Solicitante:|"
            r"Analista responsável:|"
            r"Analista responsavel:|"
            r"Descritivo da ocorrência do atendimento:|"
            r"Descritivo da ocorrencia do atendimento:|"
            r"Descritivo do atendimento:"
            r")"
        ),
        r"\n\1",
        texto
    )

    nome_empresa = (
        extrair_secao_texto(
            texto_extracao,
            [
                "Nome da empresa",
                "Empresa"
            ]
        )
    )

    empresa_loja = (
        extrair_secao_texto(
            texto_extracao,
            [
                "Empresa/Loja",
                "Loja"
            ]
        )
    )

    cnpj = (
        extrair_secao_texto(
            texto_extracao,
            [
                "CNPJ"
            ]
        )
    )

    cliente = (
        extrair_secao_texto(
            texto_extracao,
            [
                "Nome do Cliente",
                "Cliente"
            ]
        )
    )

    telefone = (
        extrair_secao_texto(
            texto_extracao,
            [
                "Telefone de contato",
                "Telefone"
            ]
        )
    )

    email = (
        extrair_secao_texto(
            texto_extracao,
            [
                "E-mail Solicitante",
                "Email",
                "E-mail"
            ]
        )
    )

    descritivo = (
        extrair_secao_texto(
            texto_extracao,
            [
                "Descritivo do atendimento",
                "Descritivo da ocorrência do atendimento",
                "Descritivo da ocorrencia do atendimento",
                "Resumo do atendimento",
                "Problema"
            ]
        )
    )

    acao = (
        extrair_secao_texto(
            texto_extracao,
            [
                "Acao realizada",
                "Ação realizada",
                "Acao realizada/orientacao",
                "Acao realizada/orientação"
            ]
        )

    )

    orientacao = (
        extrair_secao_texto(
            texto_extracao,
            [
                "Orientacao ao cliente",
                "Orientação ao cliente"
            ]
        )

    )

    resultado = (
        extrair_secao_texto(
            texto_extracao,
            [
                "Resultado"
            ]
        )

    )

    partes_descritivo = [
        parte for parte in [
            descritivo,
            acao,
            orientacao,
            resultado
        ]
        if campo_zendesk_informado(parte)
    ]

    if not partes_descritivo:

        partes_descritivo = [
            re.sub(
                r"(?:^|\n)\s*Tags\s*:.*$",
                "",
                texto_extracao,
                flags=re.IGNORECASE | re.DOTALL
            )
        ]

    descritivo_final = " ".join(partes_descritivo)

    analista = (
        extrair_secao_texto(
            texto_extracao,
            [
                "Analista responsável",
                "Analista responsavel",
                "Analista"
            ]
        )
    )

    return resumo_zendesk_exato(
        nome_empresa=nome_empresa,
        empresa_loja=empresa_loja,
        cnpj=cnpj,
        cnpj_contexto=texto_extracao,
        cliente=cliente,
        telefone=telefone,
        email=email,
        analista=analista,
        descritivo=descritivo_final
    )


def analisar_com_ia(
    transcricao,
    analista_responsavel=None,
    entidades_extraidas=None
):

    entidades_extraidas = entidades_extraidas or {}
    entidades_texto = (
        texto_entidades_extraidas(entidades_extraidas)
        or "Nenhuma entidade estruturada extraida."
    )

    prompt = f"""
Voce e um analista senior de suporte ERP.

Gere apenas um JSON valido para um atendimento de suporte ERP.
Nao gere o texto final do Zendesk.
O backend montara o texto final em ordem fixa.
Quando um campo nao aparecer na transcricao, retorne string vazia.
Nao use texto padrao de ausencia, null ou rotulos como valor.

Regras:
- Nao escrever tudo em uma linha.
- Nao inventar CNPJ, telefone, e-mail, empresa, cliente, erro, solucao ou status.
- Se nao tiver a informacao na transcricao, retorne string vazia no JSON.
- Escrever como documentacao para colar no Zendesk.
- Nao usar markdown.
- Nao usar bullets se nao houver passo a passo.
- Se houver procedimento, separar em passos numerados.
- Preserve termos tecnicos do ERP quando aparecerem.
- Use somente uma categoria principal.
- O campo descritivo deve conter apenas o texto do descritivo, sem repetir os demais campos.
- Nunca coloque rotulos de campos dentro do descritivo.
- Nunca deixe um rotulo virar valor de outro campo.
- Se o valor de um campo seria apenas "Telefone de contato:" ou "E-mail Solicitante:", retorne string vazia.
- Nunca use nome presente em saudacao ou apresentacao do analista como nome do cliente.
- Frases como "meu nome e", "sou o", "fala com" e "aqui e" normalmente indicam o analista, nao o cliente.
- Se houver duvida entre analista e cliente, deixe nome_cliente vazio.
- Se o CNPJ nao tiver exatamente 14 digitos claros, retorne vazio no JSON, exceto quando houver sequencia parecida com CNPJ.
- Nao considerar e-mail valido sem @.
- Nao preencher e-mail com dominio incompleto.
- Para CNPJ, quando houver ambiguidade, sinalizar confirmacao em vez de afirmar.
- Se a transcricao estiver confusa, curta ou cheia de ruido, documente apenas as informacoes uteis identificadas.
- Nao diga "foi identificado", "foi analisado", "foi orientado" ou "status final" se a transcricao nao mostrar isso claramente.
- Se so houver pedido de acesso remoto, registre apenas que foi solicitado acesso remoto para verificacao.
- Se nao houver conclusao explicita, finalize o descritivo com a ultima orientacao ou informacao util identificada.
- Nunca escreva frases como "O atendimento nao apresentou conclusao clara", "A transcricao esta confusa", "Nao ha informacoes suficientes", "Nao foi possivel identificar" ou "A conclusao nao ficou clara".

- Corrija termos fiscais comuns quando o contexto confirmar: ISDS-QN, ISQN ou ISS QN = ISSQN; Sintes Nacional ou Sintese Nacional = Simples Nacional; nota de servico = NFS-e; retencao de IS = retencao de ISS.
- Use correcoes de termos apenas para vocabulario tecnico. Nao use isso para inventar CNPJ, telefone, e-mail, empresa, loja ou nome de cliente.
- Nunca ignore um CNPJ parcialmente identificado.
- Se houver uma sequencia parecida com CNPJ, mas incerta, informe como "Possível CNPJ informado" e peça confirmacao.

Analista logado:
{normalizar_campo_zendesk(analista_responsavel, limite=120)}

Formato JSON:
{{
  "nome_empresa": "...",
  "empresa_loja": "...",
  "cnpj": "...",
  "nome_cliente": "...",
  "telefone": "...",
  "email": "...",
  "analista_responsavel": "...",
  "descritivo": "...",
  "sentimento_cliente": "positivo|neutro|negativo|frustrado",
  "urgencia": "baixa|media|alta|critica",
  "categoria": "fiscal|pdv|financeiro|estoque|cadastro|acesso|vendas|relatorios|integracao|outro",
  "problema_principal": "...",
  "tags": ["tag1", "tag2"]
}}

Transcricao:
{transcricao}

Entidades estruturadas extraidas pelo backend:
{entidades_texto}
"""

    resposta = cliente_resumo().chat.completions.create(
        model=SUMMARY_MODEL,
        messages=[
            {
                "role": "user",
                "content": prompt
            }
        ]
    )

    try:

        dados = (
            extrair_json_objeto(
                resposta.choices[0].message.content
            )
        )

    except Exception:

        descritivo = (
            normalizar_campo(
                resposta.choices[0].message.content,
                limite=700
            )
        )

        dados = {
            "nome_empresa": "",
            "empresa_loja": "",
            "cnpj": "",
            "nome_cliente": "",
            "telefone_contato": "",
            "email_solicitante": "",
            "analista_responsavel": analista_responsavel or "",
            "descritivo_atendimento": descritivo,
            "sentimento_cliente": "neutro",
            "urgencia": "media",
            "categoria": "outro",
            "problema_principal": descritivo[:180],
            "tags": []
        }

    descritivo = normalizar_descritivo_zendesk(
        dados.get("descritivo")
        or dados.get("descritivo_atendimento")
    )
    analista_final = (
        analista_responsavel
        or entidades_extraidas.get("analista_nome")
        or dados.get("analista_responsavel")
    )
    cliente_final = (
        entidades_extraidas.get("cliente_nome")
        or dados.get("nome_cliente")
    )

    if nomes_iguais(
        cliente_final,
        analista_final
    ) or nomes_iguais(
        cliente_final,
        entidades_extraidas.get("analista_nome")
    ):

        cliente_final = ""

    resumo = resumo_zendesk_exato(
        nome_empresa=(
            entidades_extraidas.get("empresa")
            or dados.get("nome_empresa")
        ),
        empresa_loja=dados.get("empresa_loja"),
        cnpj=(
            entidades_extraidas.get("cnpj")
            or dados.get("cnpj")
        ),
        cnpj_contexto=(
            entidades_extraidas.get("cnpj")
            or transcricao
        ),
        cliente=cliente_final,
        telefone=(
            entidades_extraidas.get("telefone")
            or dados.get("telefone")
            or dados.get("telefone_contato")
        ),
        email=(
            entidades_extraidas.get("email")
            or dados.get("email")
            or dados.get("email_solicitante")
        ),
        analista=(
            analista_final
        ),
        descritivo=descritivo
    )

    tags = (
        dados.get("tags") or []
    )

    if isinstance(tags, list):

        tags = (
            ", ".join(
                limpar_texto(tag).lower()
                for tag in tags
                if limpar_texto(tag)
            )
        )

    return {
        "resumo_zendesk": resumo[:1800],
        "sentimento_cliente": normalizar_campo(
            dados.get("sentimento_cliente"),
            "neutro",
            40
        ).lower(),
        "urgencia": normalizar_campo(
            dados.get("urgencia"),
            "media",
            40
        ).lower(),
        "categoria": normalizar_campo(
            dados.get("categoria"),
            "outro",
            80
        ).lower(),
        "problema_principal": normalizar_campo(
            dados.get("problema_principal"),
            descritivo,
            220
        ),
        "tags": normalizar_campo(
            tags,
            "",
            300
        )
    }


# =========================================
# LOGIN
# =========================================

@app.route(
    "/login",
    methods=["GET", "POST"]
)
def login():

    erro = None

    if request.method == "POST":

        usuario = limpar_texto(request.form["usuario"])
        senha = request.form["senha"]
        ip = ip_requisicao()

        if login_bloqueado(usuario, ip):

            erro = (
                "Muitas tentativas de login. "
                f"Tente novamente em {LOGIN_BLOCK_MINUTES} minutos."
            )

            return render_template(
                "login.html",
                erro=erro
            ), 429

        with conectar_banco() as conn:

            with conn.cursor() as cursor:

                cursor.execute(
                    """
                    SELECT id, senha, is_admin, ativo, perfil
                    FROM usuarios
                    WHERE usuario = %s
                    """,
                    (
                        usuario,
                    )
                )

                user = cursor.fetchone()

                senha_hash = (
                    senha_esta_em_hash(user[1])
                    if user
                    else False
                )
                senha_ok = (
                    bool(user)
                    and user[3]
                    and (
                        senha_valida(senha, user[1])
                        if senha_hash
                        else senha == user[1]
                    )
                )

                if senha_ok:

                    if not senha_hash:

                        cursor.execute(
                            """
                            UPDATE usuarios
                            SET senha = %s
                            WHERE id = %s
                            """,
                            (
                                generate_password_hash(
                                    senha
                                ),
                                user[0]
                            )
                        )

                    session["usuario_id"] = user[0]
                    perfil = (
                        user[4]
                        or (
                            "admin_tecnico"
                            if user[2]
                            else "analista"
                        )
                    )
                    session["perfil"] = perfil
                    session["is_admin"] = (
                        perfil == "admin_tecnico"
                    )
                    session["usuario_nome"] = usuario

                    limpar_tentativas_login(usuario, ip)

                    log_evento(
                        "login_sucesso",
                        usuario_id=user[0],
                        usuario=usuario,
                        perfil=perfil,
                        ip=ip
                    )

                    return redirect("/")

        registrar_login_falho(usuario, ip)

        erro = "Usuario ou senha invalidos."

    return render_template(
        "login.html",
        erro=erro
    )


@app.route("/logout")
def logout():

    session.clear()

    return redirect("/login")


@app.route("/")
def home():

    if not usuario_logado():

        return redirect("/login")

    return render_template(
        "index.html",
        is_admin=usuario_admin_tecnico(),
        is_supervisor=usuario_supervisor(),
        mostrar_custo=usuario_admin_tecnico(),
        perfil=perfil_usuario(),
        usuario_nome=session.get("usuario_nome"),
        chunk_seconds=CHUNK_SECONDS
    )


# =========================================
# ADMIN
# =========================================

@app.route(
    "/admin",
    methods=["GET", "POST"]
)
def admin_usuarios():

    if not usuario_logado():

        return redirect("/login")

    if not usuario_admin_tecnico():

        return redirect("/")

    mensagem = None
    erro = None

    if request.method == "POST":

        usuario = limpar_texto(
            request.form.get("usuario")
        )
        senha = request.form.get("senha") or ""
        perfil = request.form.get(
            "perfil",
            "analista"
        )

        if not perfil_valido(perfil):

            perfil = "analista"

        is_admin = (
            perfil == "admin_tecnico"
        )

        if not usuario or not senha:

            erro = "Informe usuario e senha."

        else:

            try:

                with conectar_banco() as conn:

                    with conn.cursor() as cursor:

                        cursor.execute(
                            """
                            INSERT INTO usuarios (
                                usuario,
                                senha,
                                is_admin,
                                perfil,
                                ativo
                            )
                            VALUES (%s, %s, %s, %s, TRUE)
                            """,
                            (
                                usuario,
                                generate_password_hash(
                                    senha
                                ),
                                is_admin,
                                perfil
                            )
                        )

                mensagem = "Usuario criado com sucesso."

            except psycopg2.errors.UniqueViolation:

                erro = "Esse usuario ja existe."

            except Exception as e:

                erro = f"Erro ao criar usuario: {e}"

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    usuario,
                    is_admin,
                    perfil,
                    ativo,
                    criado_em
                FROM usuarios
                ORDER BY usuario
                """
            )

            usuarios = cursor.fetchall()

    return render_template(
        "admin.html",
        usuarios=usuarios,
        mensagem=mensagem,
        erro=erro,
        usuario_id=usuario_logado()
    )


@app.route(
    "/admin/usuarios/<int:usuario_id>/status",
    methods=["POST"]
)
def admin_alterar_status(usuario_id):

    if not usuario_logado():

        return redirect("/login")

    if not usuario_admin_tecnico():

        return redirect("/")

    if usuario_id == usuario_logado():

        return redirect("/admin")

    ativo = request.form.get("ativo") == "true"

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                """
                UPDATE usuarios
                SET ativo = %s
                WHERE id = %s
                """,
                (
                    ativo,
                    usuario_id
                )
            )

    return redirect("/admin")


@app.route(
    "/admin/usuarios/<int:usuario_id>/senha",
    methods=["POST"]
)
def admin_alterar_senha(usuario_id):

    if not usuario_logado():

        return redirect("/login")

    if not usuario_admin_tecnico():

        return redirect("/")

    senha = request.form.get("senha") or ""

    if senha:

        with conectar_banco() as conn:

            with conn.cursor() as cursor:

                cursor.execute(
                    """
                    UPDATE usuarios
                    SET senha = %s
                    WHERE id = %s
                    """,
                    (
                        generate_password_hash(
                            senha
                        ),
                        usuario_id
                    )
                )

    return redirect("/admin")


@app.route(
    "/admin/usuarios/<int:usuario_id>/nome",
    methods=["POST"]
)
def admin_alterar_nome(usuario_id):

    if not usuario_logado():

        return redirect("/login")

    if not usuario_admin_tecnico():

        return redirect("/")

    novo_usuario = limpar_texto(
        request.form.get("usuario")
    )

    if not novo_usuario:

        return redirect("/admin")

    try:

        with conectar_banco() as conn:

            with conn.cursor() as cursor:

                cursor.execute(
                    """
                    UPDATE usuarios
                    SET usuario = %s
                    WHERE id = %s
                    """,
                    (
                        novo_usuario,
                        usuario_id
                    )
                )

        if usuario_id == usuario_logado():

            session["usuario_nome"] = novo_usuario

    except psycopg2.errors.UniqueViolation:

        pass

    return redirect("/admin")


@app.route(
    "/admin/usuarios/<int:usuario_id>/perfil",
    methods=["POST"]
)
def admin_alterar_perfil(usuario_id):

    if not usuario_logado():

        return redirect("/login")

    if not usuario_admin_tecnico():

        return redirect("/")

    perfil = request.form.get(
        "perfil",
        "analista"
    )

    if not perfil_valido(perfil):

        perfil = "analista"

    is_admin = (
        perfil == "admin_tecnico"
    )

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                """
                UPDATE usuarios
                SET
                    perfil = %s,
                    is_admin = %s
                WHERE id = %s
                """,
                (
                    perfil,
                    is_admin,
                    usuario_id
                )
            )

    if usuario_id == usuario_logado():

        session["perfil"] = perfil
        session["is_admin"] = is_admin

    return redirect("/admin")


@app.route(
    "/admin/usuarios/<int:usuario_id>/excluir",
    methods=["POST"]
)
def admin_excluir_usuario(usuario_id):

    if not usuario_logado():

        return redirect("/login")

    if not usuario_admin_tecnico():

        return redirect("/")

    if usuario_id == usuario_logado():

        return redirect("/admin")

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                """
                DELETE FROM usuarios
                WHERE id = %s
                """,
                (
                    usuario_id,
                )
            )

    return redirect("/admin")


# =========================================
# HEALTH
# =========================================

@app.route("/health")
def health():

    return jsonify({
        "status": "ok"
    })


# =========================================
# ATENDIMENTOS EM CHUNKS
# =========================================

@app.route(
    "/atendimentos/iniciar",
    methods=["POST"]
)
def iniciar_atendimento():

    usuario_id = usuario_logado()

    if not usuario_id:

        return jsonify({
            "erro": "Nao autenticado"
        }), 401

    dados = request.get_json(silent=True) or {}
    ticket_zendesk = limpar_texto(
        dados.get("ticket_zendesk", "")
    )[:80]

    data = datetime.now().strftime(
        "%d/%m/%Y %H:%M"
    )

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            uso = uso_diario_usuario(
                cursor,
                usuario_id
            )

            if uso["chamadas"] >= MAX_CALLS_PER_DAY:

                log_evento(
                    "limite_chamadas_dia",
                    usuario_id=usuario_id,
                    chamadas=uso["chamadas"],
                    limite=MAX_CALLS_PER_DAY
                )

                return erro_limite(
                    "Limite diario de atendimentos atingido.",
                    tipo="limite_chamadas_dia",
                    deve_parar_gravacao=False,
                    chamadas_hoje=uso["chamadas"],
                    limite_chamadas=MAX_CALLS_PER_DAY
                )

            if uso["segundos"] >= MAX_AUDIO_MINUTES_PER_DAY * 60:

                log_evento(
                    "limite_minutos_dia",
                    usuario_id=usuario_id,
                    segundos=uso["segundos"],
                    limite_segundos=MAX_AUDIO_MINUTES_PER_DAY * 60
                )

                return erro_limite(
                    "Limite diario de minutos de audio atingido.",
                    tipo="limite_minutos_dia",
                    deve_parar_gravacao=False,
                    minutos_hoje=round(uso["segundos"] / 60, 2),
                    limite_minutos=MAX_AUDIO_MINUTES_PER_DAY
                )

            cursor.execute(
                """
                INSERT INTO atendimentos (
                    usuario_id,
                    arquivo,
                    conteudo,
                    data,
                    status,
                    ticket_zendesk,
                    inicio_em
                )
                VALUES (%s, %s, %s, %s, %s, %s, NOW())
                RETURNING id
                """,
                (
                    usuario_id,
                    "streaming",
                    "Transcricao em andamento...",
                    data,
                    "gravando",
                    ticket_zendesk or None
                )
            )

            atendimento_id = cursor.fetchone()[0]

    log_evento(
        "atendimento_iniciado",
        usuario_id=usuario_id,
        atendimento_id=atendimento_id,
        ticket_zendesk=bool(ticket_zendesk)
    )

    return jsonify({
        "atendimento_id": atendimento_id,
        "status": "gravando"
    })


@app.route(
    "/atendimentos/chunk",
    methods=["POST"]
)
def receber_chunk():

    usuario_id = usuario_logado()

    if not usuario_id:

        return jsonify({
            "erro": "Nao autenticado"
        }), 401

    if "audio" not in request.files:

        return jsonify({
            "erro": "Sem audio"
        }), 400

    atendimento_id = request.form.get("atendimento_id")
    ordem = request.form.get("ordem")
    duracao_segundos_form = request.form.get("duracao_segundos")

    if not atendimento_id or ordem is None:

        return jsonify({
            "erro": "Atendimento ou ordem ausente"
        }), 400

    try:

        ordem_int = int(ordem)

    except ValueError:

        return jsonify({
            "erro": "Ordem invalida"
        }), 400

    try:

        duracao_chunk_segundos = int(duracao_segundos_form or CHUNK_SECONDS)

    except ValueError:

        duracao_chunk_segundos = CHUNK_SECONDS

    duracao_chunk_segundos = max(
        1,
        min(
            duracao_chunk_segundos,
            MAX_CALL_DURATION_MINUTES * 60
        )
    )

    if ordem_int >= MAX_CHUNKS_PER_CALL:

        log_evento(
            "limite_chunks_atendimento",
            usuario_id=usuario_id,
            atendimento_id=atendimento_id,
            ordem=ordem_int,
            limite=MAX_CHUNKS_PER_CALL
        )

        return erro_limite(
            "Limite de trechos por atendimento atingido.",
            tipo="limite_chunks_atendimento",
            deve_parar_gravacao=True,
            limite_chunks=MAX_CHUNKS_PER_CALL
        )

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                """
                SELECT 1
                FROM atendimentos
                WHERE id = %s
                AND usuario_id = %s
                """,
                (
                    atendimento_id,
                    usuario_id
                )
            )

            if not cursor.fetchone():

                return jsonify({
                    "erro": "Atendimento nao encontrado"
                }), 404

    arquivo = request.files["audio"]

    try:

        tamanho_audio = (
            tamanho_arquivo_upload(arquivo)
        )

        if (
            tamanho_audio < 1024
        ):

            log_evento(
                "chunk_ignorado",
                usuario_id=usuario_id,
                atendimento_id=atendimento_id,
                ordem=ordem_int,
                tamanho_audio=tamanho_audio,
                motivo="audio_muito_curto"
            )

            return jsonify({
                "status": "chunk_ignorado",
                "motivo": "audio_muito_curto"
            })

        def validar_fallback_openai():

            custo_fallback = estimar_custo_transcricao(
                duracao_chunk_segundos,
                "openai"
            )

            with conectar_banco() as conn:

                with conn.cursor() as cursor:

                    limite_resposta = (
                        validar_limite_custo_fallback_transcricao(
                            cursor,
                            usuario_id,
                            atendimento_id,
                            custo_fallback
                        )
                    )

                    if limite_resposta:

                        raise LimiteCustoFallbackTranscricao(
                            limite_resposta
                        )

            log_evento(
                "transcricao_fallback_openai",
                usuario_id=usuario_id,
                atendimento_id=atendimento_id,
                ordem=ordem_int,
                custo_estimado_usd=custo_fallback
            )

        transcricao_chunk = transcrever_chunk(
            arquivo,
            validar_fallback=validar_fallback_openai
        )

        texto_original = limpar_vazamento_prompt_transcricao(
            transcricao_chunk["texto"]
        )
        texto_normalizado = normalizar_entidades_faladas(texto_original)
        texto_limpo_para_resumo = limpar_transcricao_para_resumo(
            texto_normalizado
        )
        provider_tentado = transcricao_chunk["provider_tentado"]
        provider_usado = transcricao_chunk["provider_usado"]
        fallback_usado = transcricao_chunk["fallback_usado"]
        modelo_usado = transcricao_chunk.get("modelo_usado")

        with conectar_banco() as conn:

            with conn.cursor() as cursor:

                cursor.execute(
                    """
                    SELECT status, provider_usado, fallback_usado
                    FROM transcricoes_chunks
                    WHERE atendimento_id = %s
                    AND usuario_id = %s
                    AND ordem = %s
                    """,
                    (
                        atendimento_id,
                        usuario_id,
                        ordem_int
                    )
                )

                chunk_anterior = cursor.fetchone()

                cursor.execute(
                    """
                    INSERT INTO transcricoes_chunks (
                        atendimento_id,
                        usuario_id,
                        ordem,
                        texto,
                        status,
                        erro,
                        provider_tentado,
                        provider_usado,
                        fallback_usado,
                        duracao_segundos,
                        transcricao_bruta,
                        transcricao_normalizada,
                        transcricao_limpa_para_resumo,
                        modelo_usado,
                        tamanho_audio_original,
                        tamanho_audio_processado,
                        audio_processado,
                        audio_original_path,
                        audio_processado_path,
                        tempo_transcricao_segundos,
                        erro_preprocessamento
                    )
                    VALUES (
                        %s, %s, %s, %s, 'transcrito', NULL, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                    )
                    ON CONFLICT (atendimento_id, ordem)
                    DO UPDATE SET
                        texto = EXCLUDED.texto,
                        status = 'transcrito',
                        erro = NULL,
                        provider_tentado = EXCLUDED.provider_tentado,
                        provider_usado = EXCLUDED.provider_usado,
                        fallback_usado = EXCLUDED.fallback_usado,
                        duracao_segundos = EXCLUDED.duracao_segundos,
                        transcricao_bruta = EXCLUDED.transcricao_bruta,
                        transcricao_normalizada = EXCLUDED.transcricao_normalizada,
                        transcricao_limpa_para_resumo = EXCLUDED.transcricao_limpa_para_resumo,
                        modelo_usado = EXCLUDED.modelo_usado,
                        tamanho_audio_original = EXCLUDED.tamanho_audio_original,
                        tamanho_audio_processado = EXCLUDED.tamanho_audio_processado,
                        audio_processado = EXCLUDED.audio_processado,
                        audio_original_path = EXCLUDED.audio_original_path,
                        audio_processado_path = EXCLUDED.audio_processado_path,
                        tempo_transcricao_segundos = EXCLUDED.tempo_transcricao_segundos,
                        erro_preprocessamento = EXCLUDED.erro_preprocessamento
                    """,
                    (
                        atendimento_id,
                        usuario_id,
                        ordem_int,
                        texto_limpo_para_resumo,
                        provider_tentado,
                        provider_usado,
                        fallback_usado,
                        duracao_chunk_segundos,
                        texto_original,
                        texto_normalizado,
                        texto_limpo_para_resumo,
                        modelo_usado,
                        transcricao_chunk.get("tamanho_audio_original", tamanho_audio),
                        transcricao_chunk.get("tamanho_audio_processado", tamanho_audio),
                        bool(transcricao_chunk.get("audio_processado")),
                        transcricao_chunk.get("audio_original_path", ""),
                        transcricao_chunk.get("audio_processado_path", ""),
                        transcricao_chunk.get("tempo_transcricao_segundos", 0),
                        transcricao_chunk.get("erro_preprocessamento", "")
                    )
                )

                cursor.execute(
                    """
                    UPDATE atendimentos
                    SET status = 'transcrevendo'
                    WHERE id = %s
                    AND usuario_id = %s
                    """,
                    (
                        atendimento_id,
                        usuario_id
                    )
                )

                fallback_openai_ja_registrado = (
                    chunk_anterior
                    and chunk_anterior[0] == "transcrito"
                    and chunk_anterior[1] == "openai"
                    and chunk_anterior[2]
                )

                if (
                    fallback_usado
                    and provider_usado == "openai"
                    and not fallback_openai_ja_registrado
                ):

                    registrar_uso_evento(
                        cursor,
                        usuario_id,
                        atendimento_id,
                        "transcricao_fallback",
                        estimar_custo_transcricao(
                            duracao_chunk_segundos,
                            "openai"
                        )
                    )

        log_evento(
            "chunk_transcrito",
            usuario_id=usuario_id,
            atendimento_id=atendimento_id,
            ordem=ordem_int,
            tamanho_audio=tamanho_audio,
            tamanho_audio_original=transcricao_chunk.get("tamanho_audio_original", tamanho_audio),
            tamanho_audio_processado=transcricao_chunk.get("tamanho_audio_processado", tamanho_audio),
            audio_processado=bool(transcricao_chunk.get("audio_processado")),
            caracteres_bruto=len(texto_original),
            caracteres_limpo=len(texto_limpo_para_resumo),
            duracao_segundos=duracao_chunk_segundos,
            provider_tentado=provider_tentado,
            provider_usado=provider_usado,
            modelo_usado=modelo_usado,
            fallback_usado=fallback_usado,
            tempo_transcricao_segundos=transcricao_chunk.get("tempo_transcricao_segundos", 0),
            erro_preprocessamento=transcricao_chunk.get("erro_preprocessamento", "")
        )

        return jsonify({
            "status": "chunk_transcrito",
            "texto": texto_limpo_para_resumo,
            "provider_usado": provider_usado,
            "modelo_usado": modelo_usado,
            "fallback_usado": fallback_usado,
            "audio_processado": bool(transcricao_chunk.get("audio_processado"))
        })

    except LimiteCustoFallbackTranscricao as e:

        return e.limite_resposta

    except RateLimitError as e:

        logger.exception("LIMITE DA TRANSCRICAO ATINGIDO")

        registrar_erro_chunk(
            usuario_id,
            atendimento_id,
            ordem_int,
            e,
            status_atendimento="limite_transcricao"
        )

        if (
            TRANSCRIBE_PROVIDER == "groq"
            and TRANSCRIBE_FALLBACK_PROVIDER == "openai"
        ):

            log_evento(
                "chunk_erro_fallback",
                usuario_id=usuario_id,
                atendimento_id=atendimento_id,
                ordem=ordem_int,
                provider_tentado="groq",
                provider_usado="openai",
                fallback_usado=True,
                erro=str(e)[:500]
            )

            return jsonify({
                "erro": "Nao foi possivel transcrever este trecho. Tente novamente.",
                "status": "erro_chunk"
            }), 500

        log_evento(
            "limite_transcricao_429",
            usuario_id=usuario_id,
            atendimento_id=atendimento_id,
            ordem=ordem_int,
            erro=str(e)[:500]
        )

        return jsonify({
            "erro": "limite_atingido",
            "limite": True,
            "tipo": "limite_transcricao_groq",
            "mensagem": (
                "Limite diario da transcricao atingido. "
                "A gravacao foi pausada automaticamente. "
                "Finalize o atendimento para gerar o resumo com o conteudo ja capturado."
            ),
            "deve_parar_gravacao": True,
            "status": "limite_transcricao"
        }), 429

    except Exception as e:

        logger.exception("ERRO AO TRANSCREVER CHUNK")

        if getattr(e, "status_code", None) == 400:

            registrar_erro_chunk(
                usuario_id,
                atendimento_id,
                ordem_int,
                e
            )

            log_evento(
                "chunk_audio_invalido",
                usuario_id=usuario_id,
                atendimento_id=atendimento_id,
                ordem=ordem_int,
                provider_tentado=TRANSCRIBE_PROVIDER,
                erro=str(e)[:500]
            )

            return jsonify({
                "erro": "Audio invalido para transcricao.",
                "status": "erro_chunk"
            }), 400

        if erro_rate_limit_transcricao(e):

            registrar_erro_chunk(
                usuario_id,
                atendimento_id,
                ordem_int,
                e,
                status_atendimento="limite_transcricao"
            )

            if (
                TRANSCRIBE_PROVIDER == "groq"
                and TRANSCRIBE_FALLBACK_PROVIDER == "openai"
            ):

                log_evento(
                    "chunk_erro_fallback",
                    usuario_id=usuario_id,
                    atendimento_id=atendimento_id,
                    ordem=ordem_int,
                    provider_tentado="groq",
                    provider_usado="openai",
                    fallback_usado=True,
                    erro=str(e)[:500]
                )

                return jsonify({
                    "erro": "Nao foi possivel transcrever este trecho. Tente novamente.",
                    "status": "erro_chunk"
                }), 500

            log_evento(
                "limite_transcricao_429",
                usuario_id=usuario_id,
                atendimento_id=atendimento_id,
                ordem=ordem_int,
                erro=str(e)[:500]
            )

            return jsonify({
                "erro": "limite_atingido",
                "limite": True,
                "tipo": "limite_transcricao_groq",
                "mensagem": (
                    "Limite diario da transcricao atingido. "
                    "A gravacao foi pausada automaticamente. "
                    "Finalize o atendimento para gerar o resumo com o conteudo ja capturado."
                ),
                "deve_parar_gravacao": True,
                "status": "limite_transcricao"
            }), 429

        registrar_erro_chunk(
            usuario_id,
            atendimento_id,
            ordem_int,
            e
        )

        log_evento(
            "chunk_erro",
            usuario_id=usuario_id,
            atendimento_id=atendimento_id,
            ordem=ordem_int,
            erro=str(e)[:500]
        )

        return jsonify({
            "erro": "Nao foi possivel transcrever este trecho. Tente novamente.",
            "status": "erro_chunk"
        }), 500


def resposta_finalizacao_salva(row):

    custo_estimado = float(row[6] or 0)
    resposta = {
        "status": "finalizado",
        "resultado": texto_zendesk_formatado(row[1]),
        "chunks_total": row[2] or 0,
        "chunks_falhos": row[3] or 0,
        "chunks_ignorados": row[4] or 0,
        "segundos_transcritos": row[5] or 0,
        "reutilizado": True
    }

    if usuario_admin_tecnico():

        resposta["custo_estimado_usd"] = custo_estimado
        resposta["custo_estimado_brl"] = custo_brl(custo_estimado)

    return resposta


def conteudo_resumo_persistido(conteudo):

    if not campo_zendesk_informado(conteudo):

        return False

    comparacao = normalizar_para_comparacao(conteudo)

    return not any(
        marcador in comparacao
        for marcador in [
            "transcricao em andamento",
            "gravando",
            "processando"
        ]
    )


def liberar_finalizacao_atendimento(usuario_id, atendimento_id):

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                """
                UPDATE atendimentos
                SET status = 'transcrevendo'
                WHERE id = %s
                AND usuario_id = %s
                AND status = 'finalizando'
                """,
                (
                    atendimento_id,
                    usuario_id
                )
            )


@app.route(
    "/atendimentos/finalizar",
    methods=["POST"]
)
def finalizar_atendimento():

    usuario_id = usuario_logado()

    if not usuario_id:

        return jsonify({
            "erro": "Nao autenticado"
        }), 401

    dados = request.get_json(silent=True) or {}

    atendimento_id = dados.get("atendimento_id")
    duracao_segundos = dados.get("duracao_segundos")
    chunks_total_cliente = dados.get("chunks_total") or 0
    chunks_falhos_cliente = dados.get("chunks_falhos") or 0
    chunks_ignorados_cliente = dados.get("chunks_ignorados") or 0
    segundos_transcritos_cliente = dados.get("segundos_transcritos") or 0

    if not atendimento_id:

        return jsonify({
            "erro": "Atendimento ausente"
        }), 400

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    status,
                    conteudo,
                    chunks_total,
                    chunks_falhos,
                    chunks_ignorados,
                    segundos_transcritos,
                    custo_estimado_usd
                FROM atendimentos
                WHERE id = %s
                AND usuario_id = %s
                FOR UPDATE
                """,
                (
                    atendimento_id,
                    usuario_id
                )
            )

            atendimento_atual = cursor.fetchone()

            if not atendimento_atual:

                return jsonify({
                    "erro": "Atendimento nao encontrado"
                }), 404

            if conteudo_resumo_persistido(atendimento_atual[1]):

                log_evento(
                    "finalizacao_reutilizada",
                    usuario_id=usuario_id,
                    atendimento_id=atendimento_id
                )

                return jsonify(
                    resposta_finalizacao_salva(atendimento_atual)
                )

            if atendimento_atual[0] == "finalizando":

                log_evento(
                    "finalizacao_duplicada_ignorada",
                    usuario_id=usuario_id,
                    atendimento_id=atendimento_id
                )

                return jsonify({
                    "status": "finalizando",
                    "mensagem": "Resumo final ja esta sendo gerado."
                }), 202

            cursor.execute(
                """
                UPDATE atendimentos
                SET status = 'finalizando'
                WHERE id = %s
                AND usuario_id = %s
                """,
                (
                    atendimento_id,
                    usuario_id
                )
            )

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    COALESCE(transcricao_limpa_para_resumo, texto),
                    status,
                    COALESCE(transcricao_bruta, texto)
                FROM transcricoes_chunks
                WHERE atendimento_id = %s
                AND usuario_id = %s
                ORDER BY ordem
                """,
                (
                    atendimento_id,
                    usuario_id
                )
            )

            linhas_transcricao = cursor.fetchall()

            textos = [
                row[0]
                for row in linhas_transcricao
                if row[0]
            ]
            textos_brutos = [
                row[2]
                for row in linhas_transcricao
                if row[2]
            ]

            if not textos_brutos:

                textos_brutos = textos

            cursor.execute(
                """
                SELECT
                    COUNT(*),
                    SUM(
                        CASE
                            WHEN status = 'erro' THEN 1
                            ELSE 0
                        END
                    )
                FROM transcricoes_chunks
                WHERE atendimento_id = %s
                AND usuario_id = %s
                """,
                (
                    atendimento_id,
                    usuario_id
                )
            )

            total_banco, falhos_banco = cursor.fetchone()

            cursor.execute(
                """
                SELECT
                    COALESCE(provider_usado, %s),
                    COALESCE(duracao_segundos, %s)
                FROM transcricoes_chunks
                WHERE atendimento_id = %s
                AND usuario_id = %s
                AND status = 'transcrito'
                ORDER BY ordem
                """,
                (
                    TRANSCRIBE_PROVIDER,
                    CHUNK_SECONDS,
                    atendimento_id,
                    usuario_id
                )
            )

            chunks_transcritos_provider = [
                (
                    row[0],
                    row[1]
                )
                for row in cursor.fetchall()
            ]

            uso = uso_diario_usuario(
                cursor,
                usuario_id,
                atendimento_id
            )

    transcricao_original = limpar_vazamento_prompt_transcricao(
        limpar_texto(
            " ".join(textos_brutos)
        )
    )
    transcricao = limpar_transcricao_para_resumo(
        normalizar_entidades_faladas(
            limpar_texto(
                " ".join(textos)
            )
        )
    )
    entidades_extraidas = extrair_entidades_transcricao(transcricao_original)

    chunks_total = max(
        int(chunks_total_cliente or 0),
        int(total_banco or 0)
    )

    chunks_falhos = max(
        int(chunks_falhos_cliente or 0),
        int(falhos_banco or 0)
    )

    chunks_ignorados = (
        int(chunks_ignorados_cliente or 0)
    )

    segundos_transcritos = (
        int(segundos_transcritos_cliente or 0)
    )

    if not segundos_transcritos:

        segundos_transcritos = (
            min(
                int(duracao_segundos or 0),
                chunks_total * CHUNK_SECONDS
            )
        )

    if int(duracao_segundos or 0) > MAX_CALL_DURATION_MINUTES * 60:

        log_evento(
            "limite_duracao_atendimento",
            usuario_id=usuario_id,
            atendimento_id=atendimento_id,
            duracao_segundos=duracao_segundos,
            limite_segundos=MAX_CALL_DURATION_MINUTES * 60
        )

        liberar_finalizacao_atendimento(
            usuario_id,
            atendimento_id
        )

        return erro_limite(
            "Limite de duracao por atendimento atingido.",
            tipo="limite_duracao_atendimento",
            deve_parar_gravacao=True,
            duracao_minutos=round(int(duracao_segundos or 0) / 60, 2),
            limite_minutos=MAX_CALL_DURATION_MINUTES
        )

    if chunks_total > MAX_CHUNKS_PER_CALL:

        log_evento(
            "limite_chunks_finalizar",
            usuario_id=usuario_id,
            atendimento_id=atendimento_id,
            chunks_total=chunks_total,
            limite=MAX_CHUNKS_PER_CALL
        )

        liberar_finalizacao_atendimento(
            usuario_id,
            atendimento_id
        )

        return erro_limite(
            "Limite de trechos por atendimento atingido.",
            tipo="limite_chunks_atendimento",
            deve_parar_gravacao=True,
            chunks_total=chunks_total,
            limite_chunks=MAX_CHUNKS_PER_CALL
        )

    segundos_dia_total = (
        uso["segundos"]
        + segundos_transcritos
    )

    if segundos_dia_total > MAX_AUDIO_MINUTES_PER_DAY * 60:

        log_evento(
            "limite_minutos_finalizar_permitido",
            usuario_id=usuario_id,
            atendimento_id=atendimento_id,
            segundos_dia_total=segundos_dia_total,
            limite_segundos=MAX_AUDIO_MINUTES_PER_DAY * 60
        )

    segundos_por_provider = calcular_segundos_por_provider(
        chunks_transcritos_provider,
        segundos_transcritos
    )

    custo_estimado = estimar_custo_transcricao_por_provedor(
        segundos_por_provider
    )
    custo_transcricao_openai = estimar_custo_transcricao(
        segundos_por_provider.get("openai", 0),
        "openai"
    )

    if bool(transcricao):

        custo_estimado += estimar_custo_atendimento(0, True)
        custo_estimado = round(custo_estimado, 4)

    custo_evento_final = round(
        max(0, custo_estimado - custo_transcricao_openai),
        4
    )

    if transcricao:

        with conectar_banco() as conn:

            with conn.cursor() as cursor:

                limite_resposta = validar_limites_custo_resumo(
                    cursor,
                    usuario_id,
                    atendimento_id,
                    custo_evento_final
                )

                if limite_resposta:

                    liberar_finalizacao_atendimento(
                        usuario_id,
                        atendimento_id
                    )

                    return limite_resposta

    if transcricao:

        log_evento(
            "resumo_ia_solicitado",
            origem="finalizar_atendimento",
            usuario_id=usuario_id,
            atendimento_id=atendimento_id
        )

        try:

            analise = analisar_com_ia(
                transcricao,
                session.get("usuario_nome"),
                entidades_extraidas=entidades_extraidas
            )

        except Exception:

            liberar_finalizacao_atendimento(
                usuario_id,
                atendimento_id
            )

            raise

        resultado = analise["resumo_zendesk"]

    else:

        resultado = resumo_zendesk_exato(
            analista=session.get("usuario_nome"),
            descritivo=""
        )
        analise = {
            "sentimento_cliente": "neutro",
            "urgencia": "baixa",
            "categoria": "outro",
            "problema_principal": "Transcricao nao capturada",
            "tags": ""
        }

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                """
                UPDATE atendimentos
                SET
                    conteudo = %s,
                    transcricao_completa = %s,
                    status = 'finalizado',
                    fim_em = NOW(),
                    duracao_segundos = %s,
                    chunks_total = %s,
                    chunks_falhos = %s,
                    chunks_ignorados = %s,
                    segundos_transcritos = %s,
                    custo_estimado_usd = %s,
                    sentimento_cliente = %s,
                    urgencia = %s,
                    categoria = %s,
                    problema_principal = %s,
                    tags = %s
                WHERE id = %s
                AND usuario_id = %s
                """,
                (
                    resultado,
                    transcricao,
                    duracao_segundos,
                    chunks_total,
                    chunks_falhos,
                    chunks_ignorados,
                    segundos_transcritos,
                    custo_estimado,
                    analise["sentimento_cliente"],
                    analise["urgencia"],
                    analise["categoria"],
                    analise["problema_principal"],
                    analise["tags"],
                    atendimento_id,
                    usuario_id
                )
            )

            if transcricao:

                registrar_uso_evento(
                    cursor,
                    usuario_id,
                    atendimento_id,
                    "resumo",
                    custo_evento_final
                )

    log_evento(
        "atendimento_finalizado",
        usuario_id=usuario_id,
        atendimento_id=atendimento_id,
        chunks_total=chunks_total,
        chunks_falhos=chunks_falhos,
        chunks_ignorados=chunks_ignorados,
        segundos_transcritos=segundos_transcritos,
        custo_estimado_usd=custo_estimado,
        custo_estimado_brl=custo_brl(custo_estimado),
        segundos_por_provider=segundos_por_provider,
        custo_dia_estimado_usd=round(
            uso["custo"] + float(custo_estimado or 0),
            4
        )
    )

    resposta = {
        "status": "finalizado",
        "resultado": resultado,
        "chunks_total": chunks_total,
        "chunks_falhos": chunks_falhos,
        "chunks_ignorados": chunks_ignorados,
        "segundos_transcritos": segundos_transcritos
    }

    if usuario_admin_tecnico():

        resposta["custo_estimado_usd"] = custo_estimado
        resposta["custo_estimado_brl"] = custo_brl(custo_estimado)

    return jsonify(resposta)


# =========================================
# ROTA ANTIGA - COMPATIBILIDADE
# =========================================

@app.route(
    "/transcrever",
    methods=["POST"]
)
def transcrever_arquivo_unico():

    usuario_id = usuario_logado()

    if not usuario_id:

        return jsonify({
            "erro": "Nao autenticado"
        }), 401

    if "audio" not in request.files:

        return jsonify({
            "erro": "Sem audio"
        }), 400

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            uso = uso_diario_usuario(
                cursor,
                usuario_id
            )

            if uso["chamadas"] >= MAX_CALLS_PER_DAY:

                log_evento(
                    "limite_chamadas_upload_unico",
                    usuario_id=usuario_id,
                    chamadas=uso["chamadas"],
                    limite=MAX_CALLS_PER_DAY
                )

                return erro_limite(
                    "Limite diario de atendimentos atingido.",
                    tipo="limite_chamadas_dia",
                    deve_parar_gravacao=False,
                    chamadas_hoje=uso["chamadas"],
                    limite_chamadas=MAX_CALLS_PER_DAY
                )

            if uso["segundos"] >= MAX_AUDIO_MINUTES_PER_DAY * 60:

                log_evento(
                    "limite_minutos_upload_unico",
                    usuario_id=usuario_id,
                    segundos=uso["segundos"],
                    limite_segundos=MAX_AUDIO_MINUTES_PER_DAY * 60
                )

                return erro_limite(
                    "Limite diario de minutos de audio atingido.",
                    tipo="limite_minutos_dia",
                    deve_parar_gravacao=False,
                    minutos_hoje=round(uso["segundos"] / 60, 2),
                    limite_minutos=MAX_AUDIO_MINUTES_PER_DAY
                )

    data = datetime.now().strftime(
        "%d/%m/%Y %H:%M"
    )

    arquivo = request.files["audio"]
    custo_estimado = estimar_custo_atendimento(
        30,
        True
    )

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            limite_resposta = validar_limites_custo_resumo(
                cursor,
                usuario_id,
                None,
                custo_estimado
            )

            if limite_resposta:

                return limite_resposta

    def validar_fallback_openai_upload_unico():

        custo_fallback = estimar_custo_transcricao(
            30,
            "openai"
        )

        with conectar_banco() as conn:

            with conn.cursor() as cursor:

                limite_resposta = (
                    validar_limite_custo_fallback_transcricao(
                        cursor,
                        usuario_id,
                        None,
                        custo_fallback
                    )
                )

                if limite_resposta:

                    raise LimiteCustoFallbackTranscricao(
                        limite_resposta
                    )

        log_evento(
            "transcricao_fallback_openai",
            usuario_id=usuario_id,
            atendimento_id=None,
            ordem=None,
            custo_estimado_usd=custo_fallback
        )

    try:

        transcricao_chunk = transcrever_chunk(
            arquivo,
            validar_fallback=validar_fallback_openai_upload_unico
        )

    except LimiteCustoFallbackTranscricao as e:

        return e.limite_resposta

    texto_original = limpar_vazamento_prompt_transcricao(
        transcricao_chunk["texto"]
    )
    texto = normalizar_entidades_faladas(texto_original)
    entidades_extraidas = extrair_entidades_transcricao(texto_original)
    custo_estimado = round(
        estimar_custo_transcricao(
            30,
            transcricao_chunk["provider_usado"]
        )
        + estimar_custo_atendimento(0, True),
        4
    )

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            limite_resposta = validar_limites_custo_resumo(
                cursor,
                usuario_id,
                None,
                custo_estimado
            )

            if limite_resposta:

                return limite_resposta

    log_evento(
        "resumo_ia_solicitado",
        origem="upload_unico",
        usuario_id=usuario_id,
        atendimento_id=None
    )

    analise = analisar_com_ia(
        texto,
        session.get("usuario_nome"),
        entidades_extraidas=entidades_extraidas
    )
    resultado = analise["resumo_zendesk"]

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                """
                INSERT INTO atendimentos (
                    usuario_id,
                    arquivo,
                    conteudo,
                    data,
                    status,
                    transcricao_completa,
                    fim_em,
                    chunks_total,
                    chunks_falhos,
                    sentimento_cliente,
                    urgencia,
                    categoria,
                    problema_principal,
                    tags
                )
                VALUES (%s, %s, %s, %s, %s, %s, NOW(), %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                (
                    usuario_id,
                    str(uuid.uuid4()) + ".webm",
                    resultado,
                    data,
                    "finalizado",
                    texto,
                    1,
                    0,
                    analise["sentimento_cliente"],
                    analise["urgencia"],
                    analise["categoria"],
                    analise["problema_principal"],
                    analise["tags"]
                )
            )

            atendimento_id = cursor.fetchone()[0]

            registrar_uso_evento(
                cursor,
                usuario_id,
                atendimento_id,
                "resumo",
                custo_estimado
            )

    log_evento(
        "atendimento_upload_unico_finalizado",
        usuario_id=usuario_id,
        atendimento_id=atendimento_id,
        caracteres_transcricao=len(texto),
        categoria=analise["categoria"],
        urgencia=analise["urgencia"],
        custo_estimado_usd=custo_estimado,
        custo_estimado_brl=custo_brl(custo_estimado)
    )

    return jsonify({
        "status": "finalizado",
        "resultado": resultado
    })


# =========================================
# RESULTADOS
# =========================================

@app.route("/resultados")
def resultados():

    usuario_id = usuario_logado()

    if not usuario_id:

        return jsonify({
            "resultados": [],
            "processando": []
        })

    mostrar_custo = usuario_admin_tecnico()
    filtro_usuario = usuario_filtro_atendimentos()

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                f"""
                SELECT
                    a.id,
                    a.arquivo,
                    a.conteudo,
                    a.data,
                    a.status,
                    a.chunks_total,
                    a.chunks_falhos,
                    a.duracao_segundos,
                    a.transcricao_completa,
                    a.ticket_zendesk,
                    a.chunks_ignorados,
                    a.segundos_transcritos,
                    a.custo_estimado_usd,
                    a.resumo_editado,
                    a.sentimento_cliente,
                    a.urgencia,
                    a.categoria,
                    a.problema_principal,
                    a.tags,
                    u.usuario,
                    a.usuario_id
                FROM atendimentos a
                LEFT JOIN usuarios u
                ON u.id = a.usuario_id
                WHERE {filtro_usuario["where"]}
                ORDER BY a.id DESC
                LIMIT 300
                """,
                filtro_usuario["params"]
            )

            rows = cursor.fetchall()

            usuarios = []

            if usuario_supervisor():

                cursor.execute(
                    """
                    SELECT id, usuario
                    FROM usuarios
                    WHERE ativo = TRUE
                    ORDER BY usuario
                    """
                )

                usuarios = [
                    {
                        "id": row[0],
                        "usuario": row[1]
                    }
                    for row in cursor.fetchall()
                ]

    itens = []
    processando = []

    for row in rows:

        item = {
            "id": row[0],
            "arquivo": row[1],
            "conteudo": texto_zendesk_formatado(row[2]),
            "data": row[3],
            "status": row[4],
            "chunks_total": row[5] or 0,
            "chunks_falhos": row[6] or 0,
            "duracao_segundos": row[7] or 0,
            "transcricao_completa": row[8] or "",
            "ticket_zendesk": row[9] or "",
            "chunks_ignorados": row[10] or 0,
            "segundos_transcritos": row[11] or 0,
            "resumo_editado": bool(row[13]),
            "sentimento_cliente": row[14] or "neutro",
            "urgencia": row[15] or "media",
            "categoria": row[16] or "outro",
            "problema_principal": row[17] or "",
            "tags": row[18] or "",
            "usuario": row[19] or "",
            "usuario_id": row[20]
        }

        if mostrar_custo:

            item["custo_estimado_usd"] = float(row[12] or 0)
            item["custo_estimado_brl"] = custo_brl(
                item["custo_estimado_usd"]
            )

        itens.append(item)

        if row[4] != "finalizado":

            processando.append(str(row[0]))

    resposta = {
        "resultados": itens,
        "processando": processando,
        "escopo": filtro_usuario["escopo"],
        "analista_id": filtro_usuario["analista_id"],
        "usuarios": usuarios,
        "is_admin": usuario_admin_tecnico(),
        "is_supervisor": usuario_supervisor(),
        "mostrar_custo": mostrar_custo,
        "perfil": perfil_usuario()
    }

    if mostrar_custo:

        resposta["usd_brl_rate"] = USD_BRL_RATE

    return jsonify(resposta)


@app.route("/atendimentos/<int:atendimento_id>")
def detalhe_atendimento(atendimento_id):

    usuario_id = usuario_logado()

    if not usuario_id:

        return jsonify({
            "erro": "Nao autenticado"
        }), 401

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            if usuario_supervisor():

                cursor.execute(
                    """
                    SELECT
                        a.id,
                        a.conteudo,
                        a.transcricao_completa,
                        a.data,
                        a.status,
                        a.duracao_segundos,
                        a.chunks_total,
                        a.chunks_falhos,
                        a.ticket_zendesk,
                        a.chunks_ignorados,
                        a.segundos_transcritos,
                        a.custo_estimado_usd,
                        a.resumo_editado,
                        a.sentimento_cliente,
                        a.urgencia,
                        a.categoria,
                        a.problema_principal,
                        a.tags,
                        u.usuario
                    FROM atendimentos a
                    LEFT JOIN usuarios u
                    ON u.id = a.usuario_id
                    WHERE a.id = %s
                    """,
                    (
                        atendimento_id,
                    )
                )

            else:

                cursor.execute(
                    """
                    SELECT
                        a.id,
                        a.conteudo,
                        a.transcricao_completa,
                        a.data,
                        a.status,
                        a.duracao_segundos,
                        a.chunks_total,
                        a.chunks_falhos,
                        a.ticket_zendesk,
                        a.chunks_ignorados,
                        a.segundos_transcritos,
                        a.custo_estimado_usd,
                        a.resumo_editado,
                        a.sentimento_cliente,
                        a.urgencia,
                        a.categoria,
                        a.problema_principal,
                        a.tags,
                        u.usuario
                    FROM atendimentos a
                    LEFT JOIN usuarios u
                    ON u.id = a.usuario_id
                    WHERE a.id = %s
                    AND a.usuario_id = %s
                    """,
                    (
                        atendimento_id,
                        usuario_id
                    )
                )

            row = cursor.fetchone()

    if not row:

        return jsonify({
            "erro": "Atendimento nao encontrado"
        }), 404

    resposta = {
        "id": row[0],
        "conteudo": texto_zendesk_formatado(row[1]),
        "transcricao_completa": row[2] or "",
        "data": row[3],
        "status": row[4],
        "duracao_segundos": row[5] or 0,
        "chunks_total": row[6] or 0,
        "chunks_falhos": row[7] or 0,
        "ticket_zendesk": row[8] or "",
        "chunks_ignorados": row[9] or 0,
        "segundos_transcritos": row[10] or 0,
        "resumo_editado": bool(row[12]),
        "sentimento_cliente": row[13] or "neutro",
        "urgencia": row[14] or "media",
        "categoria": row[15] or "outro",
        "problema_principal": row[16] or "",
        "tags": row[17] or "",
        "usuario": row[18] or ""
    }

    if usuario_admin_tecnico():

        resposta["custo_estimado_usd"] = float(row[11] or 0)
        resposta["custo_estimado_brl"] = custo_brl(float(row[11] or 0))

    return jsonify(resposta)


@app.route(
    "/atendimentos/<int:atendimento_id>/resumo",
    methods=["POST"]
)
def salvar_resumo_atendimento(atendimento_id):

    usuario_id = usuario_logado()

    if not usuario_id:

        return jsonify({
            "erro": "Nao autenticado"
        }), 401

    dados = request.get_json(silent=True) or {}
    resumo = str(
        dados.get("resumo", "") or ""
    ).strip()
    resumo = texto_zendesk_formatado(
        resumo
    )
    ticket_zendesk = limpar_texto(
        dados.get("ticket_zendesk", "")
    )[:80]

    if not resumo:

        return jsonify({
            "erro": "Resumo vazio"
        }), 400

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                """
                UPDATE atendimentos
                SET
                    conteudo = %s,
                    ticket_zendesk = %s,
                    resumo_editado = TRUE,
                    resumo_editado_em = NOW()
                WHERE id = %s
                AND (
                    usuario_id = %s
                    OR %s
                )
                """,
                (
                    resumo,
                    ticket_zendesk or None,
                    atendimento_id,
                    usuario_id,
                    usuario_supervisor()
                )
            )

            if cursor.rowcount == 0:

                return jsonify({
                    "erro": "Atendimento nao encontrado"
                }), 404

    return jsonify({
        "status": "resumo_salvo",
        "resumo": resumo,
        "ticket_zendesk": ticket_zendesk
    })


@app.route(
    "/atendimentos/<int:atendimento_id>/reprocessar-resumo",
    methods=["POST"]
)
def reprocessar_resumo_atendimento(atendimento_id):

    usuario_id = usuario_logado()

    if not usuario_id:

        return jsonify({
            "erro": "Nao autenticado"
        }), 401

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    a.transcricao_completa,
                    a.segundos_transcritos,
                    a.usuario_id,
                    u.usuario
                FROM atendimentos a
                LEFT JOIN usuarios u
                ON u.id = a.usuario_id
                WHERE a.id = %s
                AND (
                    a.usuario_id = %s
                    OR %s
                )
                """,
                (
                    atendimento_id,
                    usuario_id,
                    usuario_supervisor()
                )
            )

            row = cursor.fetchone()

            if not row:

                return jsonify({
                    "erro": "Atendimento nao encontrado"
                }), 404

            transcricao = (
                limpar_vazamento_prompt_transcricao(
                    limpar_texto(row[0] or "")
                )
            )
            transcricao_original = transcricao
            transcricao = normalizar_entidades_faladas(transcricao_original)
            entidades_extraidas = extrair_entidades_transcricao(transcricao_original)

            if not transcricao:

                return jsonify({
                    "erro": "Transcricao nao disponivel"
                }), 400

            custo_estimado = estimar_custo_atendimento(
                row[1] or 0,
                True
            )
            usuario_custo_id = row[2] or usuario_id

            limite_resposta = validar_limites_custo_resumo(
                cursor,
                usuario_custo_id,
                atendimento_id,
                custo_estimado
            )

            if limite_resposta:

                return limite_resposta

            log_evento(
                "resumo_ia_solicitado",
                origem="reprocessar_resumo_manual",
                usuario_id=usuario_id,
                atendimento_id=atendimento_id
            )

            analise = analisar_com_ia(
                transcricao,
                row[3] or session.get("usuario_nome"),
                entidades_extraidas=entidades_extraidas
            )
            resumo = analise["resumo_zendesk"]

            cursor.execute(
                """
                UPDATE atendimentos
                SET
                    conteudo = %s,
                    resumo_editado = FALSE,
                    resumo_editado_em = NULL,
                    custo_estimado_usd = %s,
                    sentimento_cliente = %s,
                    urgencia = %s,
                    categoria = %s,
                    problema_principal = %s,
                    tags = %s
                WHERE id = %s
                """,
                (
                    resumo,
                    custo_estimado,
                    analise["sentimento_cliente"],
                    analise["urgencia"],
                    analise["categoria"],
                    analise["problema_principal"],
                    analise["tags"],
                    atendimento_id
                )
            )

            registrar_uso_evento(
                cursor,
                usuario_custo_id,
                atendimento_id,
                "resumo",
                custo_estimado
            )

    resposta = {
        "status": "resumo_reprocessado",
        "resumo": resumo
    }

    if usuario_admin_tecnico():

        resposta["custo_estimado_usd"] = custo_estimado
        resposta["custo_estimado_brl"] = custo_brl(custo_estimado)

    return jsonify(resposta)


@app.route(
    "/conta/senha",
    methods=["POST"]
)
def alterar_minha_senha():

    usuario_id = usuario_logado()

    if not usuario_id:

        return jsonify({
            "erro": "Nao autenticado"
        }), 401

    dados = request.get_json(silent=True) or {}
    senha_atual = dados.get("senha_atual", "")
    nova_senha = dados.get("nova_senha", "")

    if len(nova_senha) < 6:

        return jsonify({
            "erro": "A nova senha deve ter pelo menos 6 caracteres"
        }), 400

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            cursor.execute(
                """
                SELECT senha
                FROM usuarios
                WHERE id = %s
                """,
                (
                    usuario_id,
                )
            )

            row = cursor.fetchone()

            if (
                not row
                or not senha_valida(senha_atual, row[0])
            ):

                return jsonify({
                    "erro": "Senha atual invalida"
                }), 400

            cursor.execute(
                """
                UPDATE usuarios
                SET senha = %s
                WHERE id = %s
                """,
                (
                    generate_password_hash(nova_senha),
                    usuario_id
                )
            )

    return jsonify({
        "status": "senha_alterada"
    })


# =========================================
# EXPORTAR
# =========================================

@app.route("/exportar")
def exportar():

    usuario_id = usuario_logado()

    if not usuario_id:

        return redirect("/login")

    mostrar_custo = usuario_admin_tecnico()
    filtro_usuario = usuario_filtro_atendimentos()
    ids = [
        int(valor)
        for valor in re.findall(
            r"\d+",
            request.args.get("ids", "")
        )
    ]

    with conectar_banco() as conn:

        with conn.cursor() as cursor:

            where = filtro_usuario["where"]
            params = list(filtro_usuario["params"])

            if ids:

                where += " AND a.id = ANY(%s)"
                params.append(ids)

            cursor.execute(
                f"""
                SELECT
                    a.data,
                    u.usuario,
                    a.ticket_zendesk,
                    a.conteudo,
                    a.transcricao_completa,
                    a.chunks_total,
                    a.chunks_falhos,
                    a.chunks_ignorados,
                    a.segundos_transcritos,
                    a.custo_estimado_usd,
                    a.sentimento_cliente,
                    a.urgencia,
                    a.categoria,
                    a.problema_principal,
                    a.tags
                FROM atendimentos a
                LEFT JOIN usuarios u
                ON u.id = a.usuario_id
                WHERE {where}
                ORDER BY a.id DESC
                """,
                params
            )

            rows = cursor.fetchall()

    wb = Workbook()
    ws = wb.active

    ws.append([
        "Data",
        "Analista",
        "Ticket Zendesk",
        "Texto Zendesk",
        "Transcricao",
        "Trechos",
        "Trechos com falha",
        "Trechos ignorados",
        "Segundos transcritos",
        "Sentimento",
        "Urgencia",
        "Categoria",
        "Problema principal",
        "Tags internas"
    ])

    if mostrar_custo:

        ws.insert_cols(10)
        ws.cell(
            row=1,
            column=10,
            value="Custo estimado USD"
        )
        ws.insert_cols(11)
        ws.cell(
            row=1,
            column=11,
            value="Custo estimado BRL"
        )

    for row in rows:

        linha = list(row)
        linha[3] = texto_zendesk_formatado(
            linha[3]
        )

        if mostrar_custo:

            custo_estimado_usd = float(
                linha[9] or 0
            )
            linha[9] = custo_estimado_usd
            linha.insert(
                10,
                custo_brl(custo_estimado_usd)
            )
        else:

            linha.pop(9)

        ws.append(linha)

    arquivo_excel = BytesIO()
    wb.save(arquivo_excel)
    arquivo_excel.seek(0)

    return send_file(
        arquivo_excel,
        as_attachment=True,
        download_name="atendimentos.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


# =========================================
# START
# =========================================

if __name__ == "__main__":

    serve(
        app,
        host="0.0.0.0",
        port=int(
            os.getenv(
                "PORT",
                "8080"
            )
        ),
        threads=16
    )
